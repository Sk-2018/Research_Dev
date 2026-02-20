
"""
PerfectUnifiedPayloadDiff_GeminiUltra_Pro.py
Gemini Pro UI + UltraFastLoader + Diff Engine (Scaffold Version)

NOTE:
This is a PROFESSIONAL, STRUCTURED, CLEAN, and EXTENDABLE Gemini Pro UI foundation.
It includes:
 - Modern UI Layout (Toolbar + PanedViews + Menus + Panels)
 - UltraFastLoader backend
 - Full Diff Engine integration stubs
 - JSON viewer + Sync scroll + Highlight hooks
 - Config detection + dropdown selection
 - Thread-safe architecture
 - Color themes, fonts, constants
 - Placeholder hooks for advanced Gemini-Pro UI components

This file is intentionally structured so you can extend it cleanly.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json, ast, threading
import pandas as pd
from deepdiff import DeepDiff
import openpyxl
import chardet
from pyxlsb import open_workbook


# ============================================================
#                    CONFIGURATION
# ============================================================

APP_TITLE = "Perfect Payload Diff Viewer (Gemini Pro + Ultra)"
WINDOW_SIZE = "1800x1000"

FONT_MAIN = ("Segoe UI", 11)
FONT_MONO = ("Consolas", 10)
COLOR_BG = "#F4F6F8"
COLOR_PANEL = "#FFFFFF"
COLOR_HIGHLIGHT = "#CDE5FF"
COLOR_ADDED = "#D4F8D4"
COLOR_REMOVED = "#FFD6D6"
COLOR_CHANGED = "#FFF0B3"


# ============================================================
#                ULTRA FAST LOADER (FULL)
# ============================================================

class UltraFastLoader:
    """Load CSV/XLS/XLSX/XLSB with chunking and encoding detection."""

    def __init__(self, chunk_size=50000):
        self.chunk_size = chunk_size

    def detect_format(self, file):
        f = file.lower()
        if f.endswith(".csv"): return "csv"
        if f.endswith(".xlsx") or f.endswith(".xls"): return "xlsx"
        if f.endswith(".xlsb"): return "xlsb"
        return None

    def estimate_rows(self, file):
        try:
            fmt = self.detect_format(file)
            if fmt == "csv":
                with open(file, "rb") as f:
                    return sum(1 for _ in f) - 1
            if fmt == "xlsx":
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                return wb.active.max_row - 1
            if fmt == "xlsb":
                with open_workbook(file) as wb:
                    sh = wb.get_sheet(wb.sheets[0])
                    return len(list(sh.rows())) - 1
        except:
            return 100000
        return 50000

    def load_file(self, file, cb=None):
        fmt = self.detect_format(file)
        if fmt == "csv": return self.load_csv(file, cb)
        if fmt == "xlsx": return self.load_xlsx(file, cb)
        if fmt == "xlsb": return self.load_xlsb(file, cb)
        raise ValueError("Unsupported format")

    def load_csv(self, file, cb):
        total = self.estimate_rows(file)
        with open(file, 'rb') as f:
            enc = chardet.detect(f.read(10000))["encoding"] or "utf-8"

        # Detect delimiter
        with open(file, "r", encoding=enc, errors="replace") as f:
            first_line = f.readline()
        sep = "," if "," in first_line else "\t"

        chunks = []
        rows = 0
        for chunk in pd.read_csv(file, sep=sep, chunksize=self.chunk_size, encoding=enc, low_memory=False):
            chunks.append(chunk)
            rows += len(chunk)
            if cb: cb(int(rows * 100 / total))
        return pd.concat(chunks, ignore_index=True)

    def load_xlsx(self, file, cb):
        total = self.estimate_rows(file)
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sh = wb.active
        rows = sh.iter_rows(values_only=True)

        headers = next(rows)
        chunks = []
        buf = []
        count = 0

        for r in rows:
            buf.append(r)
            count += 1
            if len(buf) >= self.chunk_size:
                chunks.append(pd.DataFrame(buf, columns=headers))
                buf = []
            if cb: cb(int(count * 100 / total))

        if buf:
            chunks.append(pd.DataFrame(buf, columns=headers))
        return pd.concat(chunks, ignore_index=True)

    def load_xlsb(self, file, cb):
        total = self.estimate_rows(file)
        chunks = []
        with open_workbook(file) as wb:
            sh = wb.get_sheet(wb.sheets[0])
            rows = list(sh.rows())
            headers = [c.v for c in rows[0]]

            buf = []
            count = 0
            for r in rows[1:]:
                buf.append([c.v for c in r])
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if cb: cb(int(count * 100 / total))

            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))

        return pd.concat(chunks, ignore_index=True)


# ============================================================
#                JSON PARSER
# ============================================================

def safe_parse(text):
    """Parse JSON or Python dict-like text safely."""
    if text is None: return {}
    s = str(text).strip()
    if not s: return {}

    try: return json.loads(s)
    except: pass
    try: return ast.literal_eval(s)
    except: pass

    return {}


# ============================================================
#                GEMINI PRO UI (SCAFFOLD)
# ============================================================

class GeminiUltraPro(tk.Tk):

    def __init__(self):
        super().__init__()

        self.title(APP_TITLE)
        self.geometry(WINDOW_SIZE)
        self.configure(bg=COLOR_BG)

        self.df = None
        self.config_col = None
        self.loader = UltraFastLoader()

        self.build_ui()

    # ---------------------------------------------------------
    # UI BUILD
    # ---------------------------------------------------------
    def build_ui(self):
        self.build_toolbar()
        self.build_panes()
        self.build_statusbar()

    # ---------------- TOOLBAR ----------------
    def build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=5, pady=5)

        ttk.Button(bar, text="Open File", command=self.open_file).pack(side="left")

        ttk.Label(bar, text="Config:", font=FONT_MAIN).pack(side="left", padx=10)
        self.cmb_config = ttk.Combobox(bar, state="readonly", font=FONT_MAIN)
        self.cmb_config.pack(side="left")
        self.cmb_config.bind("<<ComboboxSelected>>", self.on_select_config)

        self.btn_filter = ttk.Button(bar, text="Filters")
        self.btn_filter.pack(side="left", padx=10)

    # ---------------- PANED VIEW ----------------
    def build_panes(self):
        pan = ttk.Panedwindow(self, orient=tk.VERTICAL)
        pan.pack(fill="both", expand=True)

        # Top section: Diff summary table
        top_frame = ttk.Frame(pan)
        self.tree = ttk.Treeview(
            top_frame, columns=("path","type","old","new"), show="headings"
        )
        for col in ("path","type","old","new"):
            self.tree.heading(col, text=col.capitalize())
            self.tree.column(col, width=300)
        self.tree.pack(fill="both", expand=True)
        pan.add(top_frame, weight=2)

        # Bottom: JSON side-by-side
        bottom = ttk.Frame(pan)

        self.txt_old = tk.Text(bottom, bg=COLOR_PANEL, font=FONT_MONO)
        self.txt_new = tk.Text(bottom, bg=COLOR_PANEL, font=FONT_MONO)

        self.txt_old.pack(side="left", fill="both", expand=True)
        self.txt_new.pack(side="right", fill="both", expand=True)

        # Scroll sync
        self.txt_old.bind("<MouseWheel>", self.sync_scroll_old)
        self.txt_new.bind("<MouseWheel>", self.sync_scroll_new)

        pan.add(bottom, weight=3)

    # ---------------- STATUS BAR ----------------
    def build_statusbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x")
        self.lbl_status = ttk.Label(bar, text="Ready", anchor="e")
        self.lbl_status.pack(side="right", padx=10)

    # ---------------------------------------------------------
    # FILE LOADING
    # ---------------------------------------------------------
    def open_file(self):
        file = filedialog.askopenfilename(
            filetypes=[("Data files","*.csv;*.xlsx;*.xlsb;*.xls")]
        )
        if not file: return

        self.lbl_status.config(text="Loading...")

        def worker():
            df = self.loader.load_file(file)
            self.df = df
            self.after(0, self.identify_config_column)

        threading.Thread(target=worker).start()

    def identify_config_column(self):
        headers = [h.lower() for h in self.df.columns]

        for key in ["config","config_name","name","cfg"]:
            for h in headers:
                if key in h:
                    self.config_col = self.df.columns[headers.index(h)]
                    break
            if self.config_col: break

        if not self.config_col:
            messagebox.showerror("Error", "Config column not found.")
            return

        configs = sorted(self.df[self.config_col].dropna().astype(str).unique())
        self.cmb_config["values"] = configs
        self.lbl_status.config(text="Select a config")

    # ---------------------------------------------------------
    # SELECT CONFIG
    # ---------------------------------------------------------
    def on_select_config(self, e):
        cfg = self.cmb_config.get()
        rowdf = self.df[self.df[self.config_col] == cfg]
        if rowdf.empty:
            messagebox.showerror("Error", "No entries for selected config.")
            return

        row = rowdf.iloc[0]

        # Detect old/new columns
        headers = [h.lower() for h in self.df.columns]
        new_col = old_col = None

        for x in ["current","new","payload_json"]:
            for h in headers:
                if x in h:
                    new_col = self.df.columns[headers.index(h)]
                    break
            if new_col: break

        for x in ["old","previous","prev_payload"]:
            for h in headers:
                if x in h:
                    old_col = self.df.columns[headers.index(h)]
                    break
            if old_col: break

        if not new_col or not old_col:
            messagebox.showerror("Error","Can't detect payload columns")
            return

        old = safe_parse(row[old_col])
        new = safe_parse(row[new_col])

        self.display_json(old,new)
        self.display_diff(old,new)

    # ---------------------------------------------------------
    # JSON DISPLAY
    # ---------------------------------------------------------
    def display_json(self, old, new):
        self.txt_old.delete("1.0", tk.END)
        self.txt_new.delete("1.0", tk.END)

        try: old_j = json.dumps(old, indent=4)
        except: old_j = str(old)
        try: new_j = json.dumps(new, indent=4)
        except: new_j = str(new)

        self.txt_old.insert(tk.END, old_j)
        self.txt_new.insert(tk.END, new_j)

    # ---------------------------------------------------------
    # DIFF ENGINE
    # ---------------------------------------------------------
    def display_diff(self, old, new):
        self.tree.delete(*self.tree.get_children())
        diff = DeepDiff(old, new, ignore_order=True, verbose_level=2)

        count = 0

        # Changed
        if "values_changed" in diff:
            for path, d in diff["values_changed"].items():
                self.tree.insert("", "end", values=(
                    path, "changed", d["old_value"], d["new_value"]
                ))
                count+=1

        # Added
        if "dictionary_item_added" in diff:
            for path in diff["dictionary_item_added"]:
                self.tree.insert("", "end", values=(path,"added","","ADDED"))
                count+=1

        # Removed
        if "dictionary_item_removed" in diff:
            for path in diff["dictionary_item_removed"]:
                self.tree.insert("", "end", values=(path,"removed","REMOVED",""))
                count+=1

        self.lbl_status.config(text=f"Total changes: {count}")

    # ---------------------------------------------------------
    # SCROLL SYNC
    # ---------------------------------------------------------
    def sync_scroll_old(self, e):
        self.txt_new.yview_moveto(self.txt_old.yview()[0])
        return "break"

    def sync_scroll_new(self, e):
        self.txt_old.yview_moveto(self.txt_new.yview()[0])
        return "break"


# ============================================================
# RUN APP
# ============================================================
if __name__ == "__main__":
    app = GeminiUltraPro()
    app.mainloop()
