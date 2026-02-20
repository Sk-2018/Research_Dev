# payload_viewer/file_loader.py
from __future__ import annotations
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import pandas as pd
from .parse_logger import ParseLogger # Import ParseLogger

# Roles that other modules import
NEEDED_ROLES = ("config_name", "config_key", "old_json", "current_json")

# Fuzzy aliases for header detection
CANON_MAP = {
    "config_name": ["config name", "configname", "cfg name", "name"],
    "config_key":  ["config key", "config_key", "confkey", "key", "id"],
    "old_json":    ["old", "old json", "old payload", "baseline", "previous", "before"],
    "current_json":["current", "current json", "current payload", "new", "after"],
}

def _norm(s: str) -> str:
    return (s or "").strip().lower().replace("_", "").replace(" ", "")

def detect_best_columns(columns: List[str]) -> Tuple[Dict[str, Optional[str]], Dict[str, float]]:
    """
    Map role -> column name (or None) using fuzzy rules on given header list.
    Returns (roles_map, confidence_map)
    """
    cols_norm = {_norm(c): c for c in columns if c is not None}
    out = {r: None for r in NEEDED_ROLES}
    confidence = {r: 0.0 for r in NEEDED_ROLES}

    # exact or alias match
    for role, aliases in CANON_MAP.items():
        for alias in aliases:
            n = _norm(alias)
            if n in cols_norm:
                out[role] = cols_norm[n]
                confidence[role] = 1.0 # Perfect match
                break
        
        # fuzzy contains if still missing
        if out[role] is None:
            seed = _norm(aliases[0])
            for k, v in cols_norm.items():
                if seed in k or k.endswith(seed):
                    out[role] = v
                    confidence[role] = 0.5 # Fuzzy match
                    break
                    
    return out, confidence

def assemble_rows(df: pd.DataFrame, roles: Dict[str, Optional[str]]) -> List[Dict[str, str]]:
    """
    Produce normalized rows with trimmed keys and raw JSON strings (no coercion).
    Keeps loader string-only; JSON is parsed later by the diff engine.
    
    'roles' is a map of {role_name: column_header_name}
    """
    def get(role_name):
        col_header = roles.get(role_name)
        if col_header and col_header in df.columns:
            return df[col_header].astype(str)
        return pd.Series([""] * len(df)) # Return empty series if role not mapped

    cn_s = get("config_name").str.strip()
    ck_s = get("config_key").str.strip()
    oj_s = get("old_json").fillna("").astype(str)
    cj_s = get("current_json").fillna("").astype(str)

    rows = []
    for a, b, o, c in zip(cn_s, ck_s, oj_s, cj_s):
        rows.append({
            "config_name": a,
            "config_key": b,
            "old_json": o,
            "current_json": c,
        })
    return rows

# --------- Excel: read header fast, then load only needed columns ---------

def _excel_header_scan(path: str, sheet: Optional[str]) -> List[str]:
    """Read only the first row via openpyxl to avoid pulling thousands of blank columns."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.worksheets[0]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    hdr = list(header)
    last_idx = 0
    for i, v in enumerate(hdr):
        if v not in (None, ""):
            last_idx = i
    return [(h if h is not None else "") for h in hdr[:last_idx + 1]]

def _guess_json_cols_from_sample(path: str, sheet: Optional[str], headers: List[str]) -> List[str]:
    """Lightweight pass to guess two JSON-like columns from first ~50 rows."""
    from openpyxl import load_workbook

    def looks_json(s: str) -> bool:
        if not isinstance(s, str):
            return False
        st = s.strip()
        return bool(st) and (st.startswith("{") or st.startswith("["))

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.worksheets[0]
    scores = [0] * len(headers)
    for row in ws.iter_rows(min_row=2, max_row=51, values_only=True):
        for i in range(min(len(headers), len(row))):
            v = row[i]
            if isinstance(v, str) and looks_json(v):
                scores[i] += 1
    ranked = sorted(range(len(headers)), key=lambda i: scores[i], reverse=True)
    out = []
    for idx in ranked:
        if scores[idx] > 0:
            out.append(headers[idx])
        if len(out) == 2:
            break
    return out

@dataclass
class FileLoader:
    parse_logger: Optional[ParseLogger] = None # <-- Accept ParseLogger
    keep_all_as_strings: bool = True
    
    def __post_init__(self):
        self.pd = pd # <-- Store pandas for app.py to check

    # ----- NEW: pre-check used by app.py -----
    def validate_file(self, path: str) -> Tuple[bool, str]:
        """
        Quick checks:
        - exists & is file
        - extension supported
        - CSV/TSV/TXT: touch small chunk
        - Excel: open first sheet and read header row
        """
        p = Path(path)
        if not p.exists():
            return False, f"File not found: {path}"
        if not p.is_file():
            return False, "Path is not a regular file"

        ext = p.suffix.lower().lstrip(".")
        allowed = {"csv", "tsv", "txt", "xlsx", "xls"}
        if ext not in allowed:
            return False, f"Unsupported file type: .{p.suffix or ''}"

        try:
            if ext in {"csv", "tsv", "txt"}:
                with p.open("rb") as f:
                    _ = f.read(2048)
            else:
                from openpyxl import load_workbook
                wb = load_workbook(filename=str(p), read_only=True, data_only=True)
                if not wb.worksheets:
                    return False, "Workbook has no sheets"
                ws = wb.worksheets[0]
                header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if header is None:
                    return False, "First sheet appears empty"
        except Exception as e:
            return False, f"Failed to read file: {e}"

        return True, ""

    # ----- FAST CSV/TSV/TXT -----
    def _pandas_load_csv_tsv_txt(self, path: str) -> pd.DataFrame:
        ext = path.lower().rsplit(".", 1)[-1]
        sep = "," if ext == "csv" else ("\t" if ext == "tsv" else None)
        df = pd.read_csv(
            path,
            sep=sep,
            dtype=str,
            keep_default_na=False,
            quoting=csv.QUOTE_MINIMAL,
            on_bad_lines="skip",
        )
        return df.applymap(lambda x: "" if x is None else str(x))

    # ----- SELECTIVE Excel read -----
    def _pandas_load_excel_selective(self, path: str, sheet: Optional[str]) -> pd.DataFrame:
        headers = _excel_header_scan(path, sheet)
        roles, confidence = detect_best_columns(headers) # Get roles to load

        target_cols: List[str] = []
        for role, col_name in roles.items():
            if col_name:
                target_cols.append(col_name)

        # If old/current not obvious, guess JSON columns from a tiny sample
        if not roles.get("old_json") or not roles.get("current_json"):
            guessed = _guess_json_cols_from_sample(path, sheet, headers)
            for g in guessed:
                if g not in target_cols:
                    target_cols.append(g)

        if not target_cols:
            # Fallback: first 10 non-empty headers
            target_cols = [h for h in headers if h][:10]

        df = pd.read_excel(
            path,
            dtype=str,
            sheet_name=sheet,
            engine="openpyxl",
            usecols=lambda c: c in set(target_cols),
        )
        df = df.fillna("").applymap(lambda x: "" if x is None else str(x))
        if not df.empty:
            df = df.loc[:, ~(df == "").all(axis=0)]  # drop fully empty columns
        return df

    # ----- PUBLIC loader -----
    def load_any(self, path: str, sheet: Optional[str] = None) -> Tuple[pd.DataFrame, Dict[str, Optional[str]], List[str], Dict[str, float]]:
        """
        Load CSV/TSV/TXT/Excel with dtype=str.
        Returns: (df, roles, problems, confidence)
        """
        problems: List[str] = []
        ext = path.lower().rsplit(".", 1)[-1]

        try:
            if ext in ("csv", "tsv", "txt"):
                df = self._pandas_load_csv_tsv_txt(path)
            else:
                df = self._pandas_load_excel_selective(path, sheet)
        except Exception as e:
            if self.parse_logger:
                self.parse_logger.log(f"Failed to load file: {e}", level="error", context=path)
            return pd.DataFrame(), {}, [f"Failed to load file: {e}"], {}

        roles, confidence = detect_best_columns(list(df.columns))

        # Primary key integrity checks
        if not roles.get("config_name"):
            problems.append("Missing 'Config Name' column.")
        if not roles.get("config_key"):
            problems.append("Missing 'Config Key' column.")
        if not roles.get("old_json") and not roles.get("current_json"):
            problems.append("No JSON payload columns detected (OLD/CURRENT).")

        if roles.get("config_name") and roles.get("config_key"):
            if roles["config_name"] in df.columns and roles["config_key"] in df.columns:
                dups = df.duplicated([roles["config_name"], roles["config_key"]]).sum()
                if dups:
                    problems.append(f"Duplicate (Config Name, Config Key) pairs: {dups}.")
                blanks_cn = (df[roles["config_name"]].str.strip() == "").sum()
                blanks_ck = (df[roles["config_key"]].str.strip() == "").sum()
                if blanks_cn or blanks_ck:
                    problems.append(f"Blank keys — Config Name: {blanks_cn}, Config Key: {blanks_ck}.")
            else:
                problems.append("Config Name or Config Key column not found in DataFrame after load.")

        # Log problems to the logger
        if self.parse_logger:
            for prob in problems:
                self.parse_logger.log(prob, level="warning", context=f"File: {os.path.basename(path)}")

        return df, roles, problems, confidence