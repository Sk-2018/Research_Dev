
# Create version with MULTIPLE selection of config keys
# Users can select multiple keys and see their diffs side-by-side or in tabs

multi_select_code = '''# -*- coding: utf-8 -*-
"""
PerfectUnifiedPayloadDiff_GeminiUltra_Pro_MULTI_SELECT.py
MULTI-SELECT VERSION: Select multiple config keys and view their diffs
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import ast
import threading
import logging
import re
from pathlib import Path
from typing import Dict, Any, Optional, List
import difflib

import pandas as pd
from deepdiff import DeepDiff
import openpyxl
import chardet

try:
    from pyxlsb import open_workbook
except ImportError:
    open_workbook = None

# Setup logging
Path('logs').mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/gemini_ultra_multi_select.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURATION
# ============================================================
APP_TITLE = "Perfect Payload Diff Viewer (Multi-Select Config Keys)"
WINDOW_SIZE = "1900x1000"
FONT_MAIN = ("Segoe UI", 10)
FONT_MONO = ("Consolas", 9)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_SMALL = ("Segoe UI", 9)

COLOR_BG = "#F4F6F8"
COLOR_PANEL = "#FFFFFF"
COLOR_HIGHLIGHT = "#87CEEB"
COLOR_ADDED = "#D4F8D4"
COLOR_REMOVED = "#FFD6D6"
COLOR_CHANGED = "#FFF0B3"
COLOR_DIFF_OLD = "#FFB6B6"
COLOR_DIFF_NEW = "#B6FFB6"
COLOR_EQUAL = "#E8E8E8"
COLOR_SELECTED = "#1976D2"

# ============================================================
# ULTRA FAST LOADER (same as before)
# ============================================================
class UltraFastLoader:
    def __init__(self, chunk_size=50000):
        self.chunk_size = chunk_size
    
    def detect_format(self, file):
        f = file.lower()
        if f.endswith((".csv", ".txt")): 
            return "csv"
        if f.endswith((".xlsx", ".xls")): 
            return "xlsx"
        if f.endswith(".xlsb"): 
            return "xlsb"
        return None
    
    def estimate_rows(self, file):
        try:
            fmt = self.detect_format(file)
            if fmt == "csv":
                with open(file, "rb") as f:
                    return sum(1 for _ in f) - 1
            if fmt == "xlsx":
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                rows = wb.active.max_row - 1
                wb.close()
                return rows
            if fmt == "xlsb" and open_workbook:
                with open_workbook(file) as wb:
                    sh = wb.get_sheet(wb.sheets[0])
                    return len(list(sh.rows())) - 1
        except Exception as e:
            logger.warning(f"Row estimation failed: {e}")
        return 50000
    
    def load_file(self, file, cb=None):
        fmt = self.detect_format(file)
        if fmt == "csv": 
            return self.load_csv(file, cb)
        if fmt == "xlsx": 
            return self.load_xlsx(file, cb)
        if fmt == "xlsb": 
            return self.load_xlsb(file, cb)
        raise ValueError(f"Unsupported format: {file}")
    
    def load_csv(self, file, cb):
        total = self.estimate_rows(file)
        with open(file, 'rb') as f:
            sample = f.read(10000)
            enc = chardet.detect(sample).get("encoding") or "utf-8"
        
        with open(file, "r", encoding=enc, errors="replace") as f:
            first_line = f.readline()
            sep = "," if "," in first_line else "\\t"
        
        chunks = []
        rows = 0
        
        try:
            for chunk in pd.read_csv(file, sep=sep, chunksize=self.chunk_size, 
                                     encoding=enc, low_memory=False, on_bad_lines='skip'):
                chunks.append(chunk)
                rows += len(chunk)
                if cb and total > 0: 
                    cb(min(100, int(rows * 100 / total)))
        except Exception as e:
            logger.error(f"CSV load error: {e}")
            if not chunks:
                raise
        
        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
    
    def load_xlsx(self, file, cb):
        total = self.estimate_rows(file)
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sh = wb.active
        rows_iter = sh.iter_rows(values_only=True)
        
        try:
            headers = next(rows_iter)
        except StopIteration:
            wb.close()
            return pd.DataFrame()
        
        chunks = []
        buf = []
        count = 0
        
        try:
            for r in rows_iter:
                buf.append(r)
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if cb and total > 0: 
                    cb(min(100, int(count * 100 / total)))
            
            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))
        finally:
            wb.close()
        
        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
    
    def load_xlsb(self, file, cb):
        if not open_workbook:
            raise ImportError("pyxlsb not installed. Install: pip install pyxlsb")
        
        total = self.estimate_rows(file)
        chunks = []
        
        with open_workbook(file) as wb:
            sh = wb.get_sheet(wb.sheets[0])
            rows = list(sh.rows())
            if not rows:
                return pd.DataFrame()
            
            headers = [c.v for c in rows[0]]
            buf = []
            count = 0
            
            for r in rows[1:]:
                buf.append([c.v for c in r])
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if cb and total > 0: 
                    cb(min(100, int(count * 100 / total)))
            
            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))
        
        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()

# ============================================================
# UTILITIES
# ============================================================
def safe_parse(text):
    if text is None: 
        return {}
    s = str(text).strip()
    if not s or s == 'nan': 
        return {}
    try: 
        return json.loads(s)
    except: 
        pass
    try: 
        return ast.literal_eval(s)
    except: 
        pass
    return {}

def pretty_json(obj):
    if obj is None: 
        return ""
    try: 
        return json.dumps(obj, indent=2, ensure_ascii=False)
    except: 
        return str(obj)

def dd_path_to_key(p: str) -> str:
    if not p:
        return p
    p = p.replace("root", "")
    p = p.replace("['", ".").replace("']", "")
    p = p.lstrip(".")
    return p

def sanitize_config_name(name: str) -> str:
    """Keep only alphanumeric and underscores"""
    if not name:
        return ""
    return re.sub(r'[^a-zA-Z0-9_]', '', str(name))

# ============================================================
# MAIN APP - MULTI-SELECT VERSION
# ============================================================
class GeminiUltraProMultiSelect(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry(WINDOW_SIZE)
        self.configure(bg=COLOR_BG)
        
        self.df = None
        self.config_name_col = None
        self.config_key_col = None
        self.old_col = None
        self.new_col = None
        self.loader = UltraFastLoader()
        self.current_config_name = None
        self.selected_config_keys = []  # Multiple keys
        
        # Two-level lookup
        self.config_name_index = {}
        self.config_data_index = {}
        
        self.build_ui()
        
        self.bind('<Control-o>', lambda e: self.open_file())
        self.bind('<F5>', lambda e: self.refresh_current())
        self.bind('<Control-a>', lambda e: self.select_all_keys())
        self.bind('<Delete>', lambda e: self.clear_selection())
        
        try:
            self.state('zoomed')
        except:
            pass
    
    def build_ui(self):
        self.build_toolbar()
        self.build_main_area()
        self.build_statusbar()
    
    def build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=10, pady=8)
        
        ttk.Button(bar, text="📁 Open File (Ctrl+O)", 
                  command=self.open_file, width=20).pack(side="left", padx=5)
        
        ttk.Label(bar, text="Config Name:", font=FONT_BOLD).pack(side="left", padx=10)
        
        self.cmb_config_name = ttk.Combobox(bar, state="readonly", font=FONT_MAIN, width=35)
        self.cmb_config_name.pack(side="left", padx=5)
        self.cmb_config_name.bind("<<ComboboxSelected>>", self.on_select_config_name)
        
        ttk.Button(bar, text="📊 Compare Selected", 
                  command=self.compare_selected, width=18).pack(side="left", padx=5)
        
        ttk.Button(bar, text="🔄 Refresh (F5)", 
                  command=self.refresh_current, width=15).pack(side="left", padx=5)
        
        self.progress = ttk.Progressbar(bar, mode='determinate', length=150)
        self.progress.pack(side="right", padx=5)
    
    def build_main_area(self):
        main_paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_paned.pack(fill="both", expand=True, padx=10, pady=5)
        
        # LEFT: Config Keys List with MULTI-SELECT
        left_frame = ttk.LabelFrame(main_paned, text="Config Keys (Multi-Select)", padding=5)
        main_paned.add(left_frame, weight=1)
        
        # Search box
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill="x", pady=(0, 5))
        ttk.Label(search_frame, text="🔍 Search:", font=FONT_SMALL).pack(side="left", padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_config_keys)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, font=FONT_MAIN)
        search_entry.pack(side="left", fill="x", expand=True, padx=5)
        
        # Action buttons
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill="x", pady=5)
        ttk.Button(btn_frame, text="✅ Select All (Ctrl+A)", 
                  command=self.select_all_keys, width=22).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="❌ Clear (Del)", 
                  command=self.clear_selection, width=15).pack(side="left", padx=2)
        
        # Listbox with EXTENDED selection mode
        list_container = ttk.Frame(left_frame)
        list_container.pack(fill="both", expand=True)
        
        self.list_config_keys = tk.Listbox(
            list_container, 
            font=FONT_MAIN, 
            selectmode=tk.EXTENDED,  # MULTI-SELECT MODE
            bg=COLOR_PANEL,
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground="#CCCCCC",
            selectbackground=COLOR_SELECTED,
            selectforeground="white"
        )
        list_scrollbar = ttk.Scrollbar(list_container, orient="vertical", 
                                       command=self.list_config_keys.yview)
        self.list_config_keys.configure(yscrollcommand=list_scrollbar.set)
        
        self.list_config_keys.pack(side="left", fill="both", expand=True)
        list_scrollbar.pack(side="right", fill="y")
        
        self.list_config_keys.bind('<<ListboxSelect>>', self.on_select_config_keys)
        
        # Count label
        self.lbl_key_count = ttk.Label(left_frame, text="0 keys | 0 selected", 
                                       font=FONT_SMALL, foreground="#666666")
        self.lbl_key_count.pack(pady=5)
        
        # RIGHT: Tabbed view for multiple diffs
        right_container = ttk.Frame(main_paned)
        main_paned.add(right_container, weight=4)
        
        self.build_tabbed_diff_viewer(right_container)
    
    def build_tabbed_diff_viewer(self, parent):
        """Build notebook with tabs for each selected key"""
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill="both", expand=True)
        
        # Info label when no selection
        self.no_selection_frame = ttk.Frame(self.notebook)
        info_label = ttk.Label(
            self.no_selection_frame, 
            text="👈 Select one or more Config Keys from the list\\n\\nTip: Use Ctrl+Click or Shift+Click for multiple selection",
            font=("Segoe UI", 12),
            foreground="#666666",
            justify=tk.CENTER
        )
        info_label.pack(expand=True)
        
        self.notebook.add(self.no_selection_frame, text="ℹ Info")
    
    def build_statusbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=10, pady=5)
        
        self.lbl_status = ttk.Label(bar, text="Ready - Load a file to begin", 
                                   anchor="w", font=FONT_MAIN)
        self.lbl_status.pack(side="left", fill="x", expand=True)
    
    def open_file(self):
        file = filedialog.askopenfilename(
            title="Select Payload File",
            filetypes=[
                ("All Supported", "*.csv *.xlsx *.xlsb *.xls *.txt"),
                ("CSV files", "*.csv *.txt"),
                ("Excel files", "*.xlsx *.xls *.xlsb"),
                ("All files", "*.*")
            ]
        )
        
        if not file:
            return
        
        self.lbl_status.config(text="Loading file...")
        self.progress['value'] = 0
        self.update_idletasks()
        
        def worker():
            try:
                def progress_cb(pct):
                    self.after(0, lambda: self.progress.configure(value=pct))
                
                df = self.loader.load_file(file, cb=progress_cb)
                self.df = df
                self.after(0, self.identify_columns)
            except Exception as e:
                logger.error(f"Load failed: {e}", exc_info=True)
                self.after(0, lambda: messagebox.showerror("Error", f"Failed to load:\\n{str(e)}"))
                self.after(0, lambda: self.lbl_status.config(text="Error loading file"))
        
        threading.Thread(target=worker, daemon=True).start()
    
    def identify_columns(self):
        """Build two-level index"""
        if self.df is None or self.df.empty:
            messagebox.showerror("Error", "No data loaded")
            return
        
        headers = [str(h).lower() if h is not None else '' for h in self.df.columns]
        
        # Find columns
        for key in ["config name", "config_name", "name", "cfg", "type"]:
            for i, h in enumerate(headers):
                if key in h:
                    self.config_name_col = self.df.columns[i]
                    break
            if self.config_name_col:
                break
        
        for key in ["config key", "config_key", "key", "id"]:
            for i, h in enumerate(headers):
                if key in h:
                    self.config_key_col = self.df.columns[i]
                    break
            if self.config_key_col:
                break
        
        for key in ["current", "new", "payload_json", "payload"]:
            for i, h in enumerate(headers):
                if key in h and "prev" not in h and "old" not in h:
                    self.new_col = self.df.columns[i]
                    break
            if self.new_col:
                break
        
        for key in ["old", "previous", "prev_payload", "prev"]:
            for i, h in enumerate(headers):
                if key in h:
                    self.old_col = self.df.columns[i]
                    break
            if self.old_col:
                break
        
        if not all([self.config_name_col, self.config_key_col, self.new_col, self.old_col]):
            messagebox.showerror("Error", "Cannot detect required columns!")
            return
        
        try:
            self.config_name_index = {}
            self.config_data_index = {}
            
            for idx, row in self.df.iterrows():
                config_name = sanitize_config_name(str(row[self.config_name_col]))
                if not config_name:
                    continue
                
                config_key = str(row[self.config_key_col])
                
                if config_name not in self.config_name_index:
                    self.config_name_index[config_name] = []
                self.config_name_index[config_name].append(config_key)
                
                composite_key = (config_name, config_key)
                self.config_data_index[composite_key] = {
                    'name': config_name,
                    'key': config_key,
                    'old': row[self.old_col],
                    'new': row[self.new_col]
                }
            
            config_names = sorted(self.config_name_index.keys())
            self.cmb_config_name["values"] = config_names
            
            total_keys = sum(len(keys) for keys in self.config_name_index.values())
            
            self.lbl_status.config(text=f"Loaded {len(self.df):,} rows - {len(config_names)} config names, {total_keys} total keys")
            self.progress['value'] = 100
            
            messagebox.showinfo("Success", 
                f"Loaded {len(self.df):,} rows\\n\\n" +
                f"Config Names: {len(config_names)}\\n" +
                f"Total Config Keys: {total_keys}\\n\\n" +
                f"💡 TIP: Select multiple keys with Ctrl+Click or Shift+Click!")
        except Exception as e:
            logger.error(f"Config processing failed: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to process:\\n{str(e)}")
    
    def on_select_config_name(self, event):
        """When config name is selected, populate the keys list"""
        config_name = self.cmb_config_name.get()
        if not config_name or config_name not in self.config_name_index:
            return
        
        self.current_config_name = config_name
        self.selected_config_keys = []
        
        self.search_var.set("")
        
        keys = sorted(self.config_name_index[config_name], key=lambda x: str(x))
        self.all_keys = keys
        
        self.list_config_keys.delete(0, tk.END)
        for key in keys:
            self.list_config_keys.insert(tk.END, key)
        
        self.update_key_count()
        self.clear_tabs()
        
        self.lbl_status.config(text=f"Config: {config_name} - Select keys to compare")
    
    def filter_config_keys(self, *args):
        """Filter config keys based on search text"""
        if not hasattr(self, 'all_keys'):
            return
        
        search_text = self.search_var.get().lower()
        
        if not search_text:
            filtered_keys = self.all_keys
        else:
            filtered_keys = [k for k in self.all_keys if search_text in str(k).lower()]
        
        self.list_config_keys.delete(0, tk.END)
        for key in filtered_keys:
            self.list_config_keys.insert(tk.END, key)
        
        self.update_key_count()
    
    def on_select_config_keys(self, event):
        """When keys are selected (multi-select)"""
        selections = self.list_config_keys.curselection()
        self.selected_config_keys = [self.list_config_keys.get(i) for i in selections]
        
        self.update_key_count()
        
        if len(self.selected_config_keys) > 0:
            self.lbl_status.config(text=f"Selected {len(self.selected_config_keys)} keys - Click 'Compare Selected' to view")
    
    def select_all_keys(self):
        """Select all visible keys"""
        self.list_config_keys.select_set(0, tk.END)
        self.on_select_config_keys(None)
    
    def clear_selection(self):
        """Clear all selections"""
        self.list_config_keys.selection_clear(0, tk.END)
        self.selected_config_keys = []
        self.update_key_count()
        self.clear_tabs()
    
    def update_key_count(self):
        """Update the key count label"""
        total = self.list_config_keys.size()
        selected = len(self.selected_config_keys)
        self.lbl_key_count.config(text=f"{total} keys | {selected} selected")
    
    def compare_selected(self):
        """Compare all selected config keys"""
        if not self.selected_config_keys:
            messagebox.showwarning("No Selection", "Please select at least one config key to compare")
            return
        
        if len(self.selected_config_keys) > 10:
            if not messagebox.askyesno("Many Selections", 
                f"You selected {len(self.selected_config_keys)} keys. This might be slow.\\nContinue?"):
                return
        
        self.clear_tabs()
        
        for key in self.selected_config_keys:
            self.add_diff_tab(key)
        
        self.lbl_status.config(text=f"Comparing {len(self.selected_config_keys)} configs")
    
    def clear_tabs(self):
        """Remove all tabs except info"""
        for tab in self.notebook.tabs():
            if self.notebook.tab(tab, "text") != "ℹ Info":
                self.notebook.forget(tab)
    
    def add_diff_tab(self, config_key):
        """Add a new tab for a config key"""
        composite_key = (self.current_config_name, config_key)
        
        if composite_key not in self.config_data_index:
            return
        
        config_data = self.config_data_index[composite_key]
        old_obj = safe_parse(config_data['old'])
        new_obj = safe_parse(config_data['new'])
        
        # Create tab
        tab_frame = ttk.Frame(self.notebook)
        self.notebook.add(tab_frame, text=str(config_key)[:20])
        
        # Build diff viewer in tab
        self.build_diff_viewer_in_tab(tab_frame, old_obj, new_obj, composite_key)
    
    def build_diff_viewer_in_tab(self, parent, old_obj, new_obj, composite_key):
        """Build complete diff viewer inside a tab"""
        paned = ttk.PanedWindow(parent, orient=tk.VERTICAL)
        paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        # TOP: Changes table
        tree_frame = ttk.LabelFrame(paned, text="Changes", padding=5)
        tree = ttk.Treeview(
            tree_frame,
            columns=("num", "path", "type", "old", "new"),
            show="headings",
            height=6
        )
        
        tree.heading("num", text="#")
        tree.heading("path", text="Path")
        tree.heading("type", text="Type")
        tree.heading("old", text="Old")
        tree.heading("new", text="New")
        
        tree.column("num", width=40, anchor="center")
        tree.column("path", width=250)
        tree.column("type", width=80, anchor="center")
        tree.column("old", width=250)
        tree.column("new", width=250)
        
        tree.tag_configure('changed', background=COLOR_CHANGED)
        tree.tag_configure('added', background=COLOR_ADDED)
        tree.tag_configure('removed', background=COLOR_REMOVED)
        
        tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=tree_scroll.set)
        tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")
        
        paned.add(tree_frame, weight=1)
        
        # BOTTOM: JSON viewers
        json_frame = ttk.LabelFrame(paned, text="Payloads", padding=5)
        json_paned = ttk.PanedWindow(json_frame, orient=tk.HORIZONTAL)
        json_paned.pack(fill="both", expand=True)
        
        # Old
        old_frame = ttk.Frame(json_paned)
        json_paned.add(old_frame, weight=1)
        ttk.Label(old_frame, text="OLD", font=FONT_BOLD, foreground='#C62828').pack()
        txt_old = tk.Text(old_frame, wrap="none", font=FONT_MONO, bg=COLOR_PANEL, height=15)
        old_scroll = ttk.Scrollbar(old_frame, command=txt_old.yview)
        txt_old.configure(yscrollcommand=old_scroll.set)
        txt_old.pack(side="left", fill="both", expand=True)
        old_scroll.pack(side="right", fill="y")
        txt_old.insert("1.0", pretty_json(old_obj))
        txt_old.configure(state="disabled")
        
        # New
        new_frame = ttk.Frame(json_paned)
        json_paned.add(new_frame, weight=1)
        ttk.Label(new_frame, text="CURRENT", font=FONT_BOLD, foreground='#2E7D32').pack()
        txt_new = tk.Text(new_frame, wrap="none", font=FONT_MONO, bg=COLOR_PANEL, height=15)
        new_scroll = ttk.Scrollbar(new_frame, command=txt_new.yview)
        txt_new.configure(yscrollcommand=new_scroll.set)
        txt_new.pack(side="left", fill="both", expand=True)
        new_scroll.pack(side="right", fill="y")
        txt_new.insert("1.0", pretty_json(new_obj))
        txt_new.configure(state="disabled")
        
        paned.add(json_frame, weight=2)
        
        # Populate diff
        diff = DeepDiff(old_obj, new_obj, ignore_order=True, verbose_level=2)
        count = 0
        
        for path, change in diff.get("values_changed", {}).items():
            count += 1
            tree.insert("", "end", 
                       values=(count, dd_path_to_key(path), "CHANGED", 
                              str(change.get('old_value'))[:80], 
                              str(change.get('new_value'))[:80]),
                       tags=('changed',))
        
        for path, change in diff.get("type_changes", {}).items():
            count += 1
            tree.insert("", "end",
                       values=(count, dd_path_to_key(path), "TYPE CHG",
                              str(change.get('old_value'))[:80],
                              str(change.get('new_value'))[:80]),
                       tags=('changed',))
        
        for path in diff.get("dictionary_item_added", set()):
            count += 1
            tree.insert("", "end",
                       values=(count, dd_path_to_key(path), "ADDED", "", "ADDED"),
                       tags=('added',))
        
        for path in diff.get("dictionary_item_removed", set()):
            count += 1
            tree.insert("", "end",
                       values=(count, dd_path_to_key(path), "REMOVED", "REMOVED", ""),
                       tags=('removed',))
    
    def refresh_current(self):
        if self.selected_config_keys:
            self.compare_selected()

if __name__ == "__main__":
    try:
        app = GeminiUltraProMultiSelect()
        app.mainloop()
    except Exception as e:
        print(f"ERROR: {e}")
        logger.error(f"App failed: {e}", exc_info=True)
        input("Press Enter to exit...")
'''

# Save the multi-select code
with open("PerfectUnifiedPayloadDiff_GeminiUltra_Pro_MULTI_SELECT.py", "w", encoding="utf-8") as f:
    f.write(multi_select_code)

print("✅ MULTI-SELECT VERSION SAVED!")
print("\nFile: PerfectUnifiedPayloadDiff_GeminiUltra_Pro_MULTI_SELECT.py")
print("\n" + "="*80)
print("MULTI-SELECT FEATURES:")
print("="*80)
print("1. ✅ Extended Selection Mode: Click, Ctrl+Click, Shift+Click")
print("2. ✅ Select All Button: Select all visible keys (Ctrl+A)")
print("3. ✅ Clear Button: Deselect all (Delete key)")
print("4. ✅ Compare Button: Generate tabs for all selected keys")
print("5. ✅ Tabbed Interface: Each key gets its own tab")
print("6. ✅ Selection Counter: Shows 'X keys | Y selected'")
print("7. ✅ Smart Warning: Warns if selecting >10 keys")
print("\n" + "="*80)
print("USAGE EXAMPLES:")
print("="*80)
print("• Select Single: Click on a key")
print("• Select Multiple: Ctrl+Click multiple keys")
print("• Select Range: Click first, Shift+Click last")
print("• Select All: Click 'Select All' or press Ctrl+A")
print("• Compare: Click 'Compare Selected' button")
print("• Navigate: Switch between tabs to see each diff")
print("\n✨ Perfect for comparing multiple config keys at once!")
