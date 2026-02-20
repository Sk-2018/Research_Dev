# -*- coding: utf-8 -*-
"""
PerfectUnifiedPayloadDiff_GeminiUltra_Pro_ENHANCED_V2.py
ENHANCED VERSION: Two-stage selection with Config Name dropdown + Config Key list
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import ast
import threading
import logging
import re
from pathlib import Path
from typing import Dict, Any, Optional, List
import difflib

import pandas as pd
from deepdiff import DeepDiff
import openpyxl
import chardet

try:
    from pyxlsb import open_workbook
except ImportError:
    open_workbook = None

# Setup logging
Path('logs').mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/gemini_ultra_enhanced_v2.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# CONFIGURATION
# ============================================================
APP_TITLE = "Perfect Payload Diff Viewer (Enhanced Two-Stage Selection)"
WINDOW_SIZE = "1900x1000"
FONT_MAIN = ("Segoe UI", 10)
FONT_MONO = ("Consolas", 9)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_SMALL = ("Segoe UI", 9)

COLOR_BG = "#F4F6F8"
COLOR_PANEL = "#FFFFFF"
COLOR_HIGHLIGHT = "#87CEEB"
COLOR_ADDED = "#D4F8D4"
COLOR_REMOVED = "#FFD6D6"
COLOR_CHANGED = "#FFF0B3"
COLOR_DIFF_OLD = "#FFB6B6"
COLOR_DIFF_NEW = "#B6FFB6"
COLOR_EQUAL = "#E8E8E8"
COLOR_SELECTED = "#E3F2FD"

# ============================================================
# ULTRA FAST LOADER
# ============================================================
class UltraFastLoader:
    def __init__(self, chunk_size=50000):
        self.chunk_size = chunk_size

    def detect_format(self, file):
        f = file.lower()
        if f.endswith((".csv", ".txt")): 
            return "csv"
        if f.endswith((".xlsx", ".xls")): 
            return "xlsx"
        if f.endswith(".xlsb"): 
            return "xlsb"
        return None

    def estimate_rows(self, file):
        try:
            fmt = self.detect_format(file)
            if fmt == "csv":
                with open(file, "rb") as f:
                    return sum(1 for _ in f) - 1
            if fmt == "xlsx":
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                rows = wb.active.max_row - 1
                wb.close()
                return rows
            if fmt == "xlsb" and open_workbook:
                with open_workbook(file) as wb:
                    sh = wb.get_sheet(wb.sheets[0])
                    return len(list(sh.rows())) - 1
        except Exception as e:
            logger.warning(f"Row estimation failed: {e}")
        return 50000

    def load_file(self, file, cb=None):
        fmt = self.detect_format(file)
        if fmt == "csv": 
            return self.load_csv(file, cb)
        if fmt == "xlsx": 
            return self.load_xlsx(file, cb)
        if fmt == "xlsb": 
            return self.load_xlsb(file, cb)
        raise ValueError(f"Unsupported format: {file}")

    def load_csv(self, file, cb):
        total = self.estimate_rows(file)
        with open(file, 'rb') as f:
            sample = f.read(10000)
            enc = chardet.detect(sample).get("encoding") or "utf-8"

        with open(file, "r", encoding=enc, errors="replace") as f:
            first_line = f.readline()
            sep = "," if "," in first_line else "\t"

        chunks = []
        rows = 0

        try:
            for chunk in pd.read_csv(file, sep=sep, chunksize=self.chunk_size, 
                                     encoding=enc, low_memory=False, on_bad_lines='skip'):
                chunks.append(chunk)
                rows += len(chunk)
                if cb and total > 0: 
                    cb(min(100, int(rows * 100 / total)))
        except Exception as e:
            logger.error(f"CSV load error: {e}")
            if not chunks:
                raise

        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()

    def load_xlsx(self, file, cb):
        total = self.estimate_rows(file)
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sh = wb.active
        rows_iter = sh.iter_rows(values_only=True)

        try:
            headers = next(rows_iter)
        except StopIteration:
            wb.close()
            return pd.DataFrame()

        chunks = []
        buf = []
        count = 0

        try:
            for r in rows_iter:
                buf.append(r)
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if cb and total > 0: 
                    cb(min(100, int(count * 100 / total)))

            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))
        finally:
            wb.close()

        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()

    def load_xlsb(self, file, cb):
        if not open_workbook:
            raise ImportError("pyxlsb not installed. Install: pip install pyxlsb")

        total = self.estimate_rows(file)
        chunks = []

        with open_workbook(file) as wb:
            sh = wb.get_sheet(wb.sheets[0])
            rows = list(sh.rows())
            if not rows:
                return pd.DataFrame()

            headers = [c.v for c in rows[0]]
            buf = []
            count = 0

            for r in rows[1:]:
                buf.append([c.v for c in r])
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if cb and total > 0: 
                    cb(min(100, int(count * 100 / total)))

            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))

        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()

# ============================================================
# UTILITIES
# ============================================================
def safe_parse(text):
    if text is None: 
        return {}
    s = str(text).strip()
    if not s or s == 'nan': 
        return {}
    try: 
        return json.loads(s)
    except: 
        pass
    try: 
        return ast.literal_eval(s)
    except: 
        pass
    return {}

def pretty_json(obj):
    if obj is None: 
        return ""
    try: 
        return json.dumps(obj, indent=2, ensure_ascii=False)
    except: 
        return str(obj)

def dd_path_to_key(p: str) -> str:
    if not p:
        return p
    p = p.replace("root", "")
    p = p.replace("['", ".").replace("']", "")
    p = p.lstrip(".")
    return p

def sanitize_config_name(name: str) -> str:
    """Keep only alphanumeric and underscores"""
    if not name:
        return ""
    # Keep only a-z, A-Z, 0-9, and underscore
    return re.sub(r'[^a-zA-Z0-9_]', '', str(name))

# ============================================================
# MAIN APP - ENHANCED TWO-STAGE SELECTION
# ============================================================
class GeminiUltraProEnhanced(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry(WINDOW_SIZE)
        self.configure(bg=COLOR_BG)

        self.df = None
        self.config_name_col = None
        self.config_key_col = None
        self.old_col = None
        self.new_col = None
        self.loader = UltraFastLoader()
        self.current_config_name = None
        self.current_config_key = None
        self.config_diffs = {}

        # OPTIMIZATION: Two-level lookup
        self.config_name_index = {}  # config_name -> list of config_keys
        self.config_data_index = {}  # (config_name, config_key) -> row_data

        self.build_ui()

        self.bind('<Control-o>', lambda e: self.open_file())
        self.bind('<F5>', lambda e: self.refresh_current())

        try:
            self.state('zoomed')
        except:
            pass

    def build_ui(self):
        self.build_toolbar()
        self.build_main_area()
        self.build_statusbar()

    def build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=10, pady=8)

        ttk.Button(bar, text="📁 Open File (Ctrl+O)", 
                  command=self.open_file, width=20).pack(side="left", padx=5)

        ttk.Label(bar, text="Config Name:", font=FONT_BOLD).pack(side="left", padx=10)

        # Config Name Dropdown
        self.cmb_config_name = ttk.Combobox(bar, state="readonly", font=FONT_MAIN, width=40)
        self.cmb_config_name.pack(side="left", padx=5)
        self.cmb_config_name.bind("<<ComboboxSelected>>", self.on_select_config_name)

        ttk.Button(bar, text="🔄 Refresh (F5)", 
                  command=self.refresh_current, width=15).pack(side="left", padx=5)

        self.progress = ttk.Progressbar(bar, mode='determinate', length=150)
        self.progress.pack(side="right", padx=5)

    def build_main_area(self):
        # Main horizontal split: Config Keys List | Diff Viewers
        main_paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_paned.pack(fill="both", expand=True, padx=10, pady=5)

        # LEFT: Config Keys List
        left_frame = ttk.LabelFrame(main_paned, text="Config Keys", padding=5)
        main_paned.add(left_frame, weight=1)

        # Search box for config keys
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill="x", pady=(0, 5))
        ttk.Label(search_frame, text="🔍 Search:", font=FONT_SMALL).pack(side="left", padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_config_keys)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, font=FONT_MAIN)
        search_entry.pack(side="left", fill="x", expand=True, padx=5)

        # Listbox for config keys
        list_container = ttk.Frame(left_frame)
        list_container.pack(fill="both", expand=True)

        self.list_config_keys = tk.Listbox(
            list_container, 
            font=FONT_MAIN, 
            selectmode=tk.SINGLE,
            bg=COLOR_PANEL,
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground="#CCCCCC"
        )
        list_scrollbar = ttk.Scrollbar(list_container, orient="vertical", 
                                       command=self.list_config_keys.yview)
        self.list_config_keys.configure(yscrollcommand=list_scrollbar.set)

        self.list_config_keys.pack(side="left", fill="both", expand=True)
        list_scrollbar.pack(side="right", fill="y")

        self.list_config_keys.bind('<<ListboxSelect>>', self.on_select_config_key)

        # Count label
        self.lbl_key_count = ttk.Label(left_frame, text="0 keys", font=FONT_SMALL, 
                                       foreground="#666666")
        self.lbl_key_count.pack(pady=5)

        # RIGHT: Diff Viewers (existing panes)
        right_container = ttk.Frame(main_paned)
        main_paned.add(right_container, weight=4)

        self.build_diff_panes(right_container)

    def build_diff_panes(self, parent):
        """Build the diff viewer panes (same as before)"""
        diff_paned = ttk.PanedWindow(parent, orient=tk.VERTICAL)
        diff_paned.pack(fill="both", expand=True)

        # TOP: Diff table
        top_frame = ttk.LabelFrame(diff_paned, text="Changes Summary", padding=5)

        tree_container = ttk.Frame(top_frame)
        tree_container.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(
            tree_container,
            columns=("num", "path", "type", "old", "new"),
            show="headings",
            height=8
        )

        self.tree.heading("num", text="#")
        self.tree.heading("path", text="JSON Path")
        self.tree.heading("type", text="Type")
        self.tree.heading("old", text="Old Value")
        self.tree.heading("new", text="New Value")

        self.tree.column("num", width=50, anchor="center")
        self.tree.column("path", width=300)
        self.tree.column("type", width=100, anchor="center")
        self.tree.column("old", width=300)
        self.tree.column("new", width=300)

        self.tree.tag_configure('changed', background=COLOR_CHANGED)
        self.tree.tag_configure('added', background=COLOR_ADDED)
        self.tree.tag_configure('removed', background=COLOR_REMOVED)

        tree_scroll = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        self.tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")

        diff_paned.add(top_frame, weight=1)

        # MIDDLE: Inline diff
        inline_frame = ttk.LabelFrame(diff_paned, text="Selected Item - Character Diff", padding=5)

        inline_paned = ttk.PanedWindow(inline_frame, orient=tk.HORIZONTAL)
        inline_paned.pack(fill="both", expand=True)

        # Old
        old_inline = ttk.Frame(inline_paned)
        inline_paned.add(old_inline, weight=1)
        ttk.Label(old_inline, text="Old Value", font=FONT_BOLD, foreground='#C62828').pack()

        self.txt_inline_old = tk.Text(old_inline, height=5, wrap='word', font=FONT_MONO, bg=COLOR_PANEL)
        old_inline_scroll = ttk.Scrollbar(old_inline, command=self.txt_inline_old.yview)
        self.txt_inline_old.configure(yscrollcommand=old_inline_scroll.set)
        self.txt_inline_old.pack(side="left", fill="both", expand=True)
        old_inline_scroll.pack(side="right", fill="y")

        self.txt_inline_old.tag_configure('equal', background=COLOR_EQUAL)
        self.txt_inline_old.tag_configure('delete', background=COLOR_DIFF_OLD, foreground='black')

        # New
        new_inline = ttk.Frame(inline_paned)
        inline_paned.add(new_inline, weight=1)
        ttk.Label(new_inline, text="New Value", font=FONT_BOLD, foreground='#2E7D32').pack()

        self.txt_inline_new = tk.Text(new_inline, height=5, wrap='word', font=FONT_MONO, bg=COLOR_PANEL)
        new_inline_scroll = ttk.Scrollbar(new_inline, command=self.txt_inline_new.yview)
        self.txt_inline_new.configure(yscrollcommand=new_inline_scroll.set)
        self.txt_inline_new.pack(side="left", fill="both", expand=True)
        new_inline_scroll.pack(side="right", fill="y")

        self.txt_inline_new.tag_configure('equal', background=COLOR_EQUAL)
        self.txt_inline_new.tag_configure('insert', background=COLOR_DIFF_NEW, foreground='black')

        diff_paned.add(inline_frame, weight=1)

        # BOTTOM: Full JSON with sync scroll
        payload_frame = ttk.LabelFrame(diff_paned, text="Full Payloads - SYNCHRONIZED", padding=5)

        payload_paned = ttk.PanedWindow(payload_frame, orient=tk.HORIZONTAL)
        payload_paned.pack(fill="both", expand=True)

        # OLD
        old_payload = ttk.Frame(payload_paned)
        payload_paned.add(old_payload, weight=1)
        ttk.Label(old_payload, text="OLD Payload", font=FONT_BOLD, foreground='#C62828').pack()

        old_container = ttk.Frame(old_payload)
        old_container.pack(fill="both", expand=True)

        self.txt_old = tk.Text(old_container, wrap="none", font=FONT_MONO, bg=COLOR_PANEL)
        self.old_scroll_y = ttk.Scrollbar(old_container, orient="vertical", command=self._on_old_scroll)
        self.old_scroll_x = ttk.Scrollbar(old_container, orient="horizontal", command=self.txt_old.xview)
        self.txt_old.configure(yscrollcommand=self._update_old_scrollbar, xscrollcommand=self.old_scroll_x.set)

        self.txt_old.grid(row=0, column=0, sticky="nsew")
        self.old_scroll_y.grid(row=0, column=1, sticky="ns")
        self.old_scroll_x.grid(row=1, column=0, sticky="ew")

        old_container.grid_rowconfigure(0, weight=1)
        old_container.grid_columnconfigure(0, weight=1)

        self.txt_old.tag_configure('highlight', background=COLOR_HIGHLIGHT, foreground='black')

        # CURRENT
        new_payload = ttk.Frame(payload_paned)
        payload_paned.add(new_payload, weight=1)
        ttk.Label(new_payload, text="CURRENT Payload", font=FONT_BOLD, foreground='#2E7D32').pack()

        new_container = ttk.Frame(new_payload)
        new_container.pack(fill="both", expand=True)

        self.txt_new = tk.Text(new_container, wrap="none", font=FONT_MONO, bg=COLOR_PANEL)
        self.new_scroll_y = ttk.Scrollbar(new_container, orient="vertical", command=self._on_new_scroll)
        self.new_scroll_x = ttk.Scrollbar(new_container, orient="horizontal", command=self.txt_new.xview)
        self.txt_new.configure(yscrollcommand=self._update_new_scrollbar, xscrollcommand=self.new_scroll_x.set)

        self.txt_new.grid(row=0, column=0, sticky="nsew")
        self.new_scroll_y.grid(row=0, column=1, sticky="ns")
        self.new_scroll_x.grid(row=1, column=0, sticky="ew")

        new_container.grid_rowconfigure(0, weight=1)
        new_container.grid_columnconfigure(0, weight=1)

        self.txt_new.tag_configure('highlight', background=COLOR_HIGHLIGHT, foreground='black')

        # Sync mousewheel
        self.txt_old.bind('<MouseWheel>', self._on_mousewheel_old)
        self.txt_new.bind('<MouseWheel>', self._on_mousewheel_new)
        self.txt_old.bind('<Button-4>', self._on_mousewheel_old)
        self.txt_old.bind('<Button-5>', self._on_mousewheel_old)
        self.txt_new.bind('<Button-4>', self._on_mousewheel_new)
        self.txt_new.bind('<Button-5>', self._on_mousewheel_new)

        diff_paned.add(payload_frame, weight=2)

        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)

    def build_statusbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=10, pady=5)

        self.lbl_status = ttk.Label(bar, text="Ready - Load a file to begin", 
                                   anchor="w", font=FONT_MAIN)
        self.lbl_status.pack(side="left", fill="x", expand=True)

    # Synchronized scrolling
    def _on_old_scroll(self, *args):
        self.txt_old.yview(*args)
        self.txt_new.yview(*args)

    def _on_new_scroll(self, *args):
        self.txt_new.yview(*args)
        self.txt_old.yview(*args)

    def _update_old_scrollbar(self, first, last):
        self.old_scroll_y.set(first, last)
        if self.txt_new.yview() != (float(first), float(last)):
            self.txt_new.yview_moveto(first)

    def _update_new_scrollbar(self, first, last):
        self.new_scroll_y.set(first, last)
        if self.txt_old.yview() != (float(first), float(last)):
            self.txt_old.yview_moveto(first)

    def _on_mousewheel_old(self, event):
        delta = -1 if event.delta < 0 or event.num == 5 else 1
        self.txt_old.yview_scroll(delta, "units")
        self.txt_new.yview_scroll(delta, "units")
        return "break"

    def _on_mousewheel_new(self, event):
        delta = -1 if event.delta < 0 or event.num == 5 else 1
        self.txt_new.yview_scroll(delta, "units")
        self.txt_old.yview_scroll(delta, "units")
        return "break"

    # File loading
    def open_file(self):
        file = filedialog.askopenfilename(
            title="Select Payload File",
            filetypes=[
                ("All Supported", "*.csv *.xlsx *.xlsb *.xls *.txt"),
                ("CSV files", "*.csv *.txt"),
                ("Excel files", "*.xlsx *.xls *.xlsb"),
                ("All files", "*.*")
            ]
        )

        if not file:
            return

        self.lbl_status.config(text="Loading file...")
        self.progress['value'] = 0
        self.update_idletasks()

        def worker():
            try:
                def progress_cb(pct):
                    self.after(0, lambda: self.progress.configure(value=pct))

                df = self.loader.load_file(file, cb=progress_cb)
                self.df = df
                self.after(0, self.identify_columns)
            except Exception as e:
                logger.error(f"Load failed: {e}", exc_info=True)
                self.after(0, lambda: messagebox.showerror("Error", f"Failed to load:\n{str(e)}"))
                self.after(0, lambda: self.lbl_status.config(text="Error loading file"))

        threading.Thread(target=worker, daemon=True).start()

    def identify_columns(self):
        """ENHANCED: Build two-level index"""
        if self.df is None or self.df.empty:
            messagebox.showerror("Error", "No data loaded")
            return

        headers = [str(h).lower() if h is not None else '' for h in self.df.columns]

        # Find config name
        for key in ["config name", "config_name", "name", "cfg", "type"]:
            for i, h in enumerate(headers):
                if key in h:
                    self.config_name_col = self.df.columns[i]
                    break
            if self.config_name_col:
                break

        # Find config key
        for key in ["config key", "config_key", "key", "id"]:
            for i, h in enumerate(headers):
                if key in h:
                    self.config_key_col = self.df.columns[i]
                    break
            if self.config_key_col:
                break

        # Find new/current
        for key in ["current", "new", "payload_json", "payload"]:
            for i, h in enumerate(headers):
                if key in h and "prev" not in h and "old" not in h:
                    self.new_col = self.df.columns[i]
                    break
            if self.new_col:
                break

        # Find old/previous
        for key in ["old", "previous", "prev_payload", "prev"]:
            for i, h in enumerate(headers):
                if key in h:
                    self.old_col = self.df.columns[i]
                    break
            if self.old_col:
                break

        if not self.config_name_col or not self.config_key_col or not self.new_col or not self.old_col:
            messagebox.showerror("Error", 
                f"Cannot detect columns!\n\n" +
                f"Config Name: {self.config_name_col}\n" +
                f"Config Key: {self.config_key_col}\n" +
                f"Current: {self.new_col}\n" +
                f"Previous: {self.old_col}\n\n" +
                f"Available: {', '.join([str(c) for c in self.df.columns])}")
            return

        try:
            # Build two-level index
            self.config_name_index = {}
            self.config_data_index = {}

            for idx, row in self.df.iterrows():
                # Sanitize config name
                raw_config_name = str(row[self.config_name_col])
                config_name = sanitize_config_name(raw_config_name)

                if not config_name:
                    continue

                config_key = str(row[self.config_key_col])

                # Build name -> keys mapping
                if config_name not in self.config_name_index:
                    self.config_name_index[config_name] = []
                self.config_name_index[config_name].append(config_key)

                # Build (name, key) -> data mapping
                composite_key = (config_name, config_key)
                self.config_data_index[composite_key] = {
                    'name': config_name,
                    'key': config_key,
                    'old': row[self.old_col],
                    'new': row[self.new_col]
                }

            # Sort config names
            config_names = sorted(self.config_name_index.keys())
            self.cmb_config_name["values"] = config_names

            total_keys = sum(len(keys) for keys in self.config_name_index.values())

            self.lbl_status.config(text=f"Loaded {len(self.df):,} rows - {len(config_names)} config names, {total_keys} total keys")
            self.progress['value'] = 100

            messagebox.showinfo("Success", 
                f"Loaded {len(self.df):,} rows\n\n" +
                f"Config Names: {len(config_names)}\n" +
                f"Total Config Keys: {total_keys}\n\n" +
                f"Select a Config Name to see its keys!")
        except Exception as e:
            logger.error(f"Config processing failed: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to process:\n{str(e)}")

    def on_select_config_name(self, event):
        """When config name is selected, populate the keys list"""
        config_name = self.cmb_config_name.get()
        if not config_name or config_name not in self.config_name_index:
            return

        self.current_config_name = config_name
        self.current_config_key = None

        # Clear search
        self.search_var.set("")

        # Populate keys list
        keys = sorted(self.config_name_index[config_name], key=lambda x: str(x))
        self.all_keys = keys  # Store for filtering

        self.list_config_keys.delete(0, tk.END)
        for key in keys:
            self.list_config_keys.insert(tk.END, key)

        self.lbl_key_count.config(text=f"{len(keys)} keys")

        # Clear diff views
        self.clear_diff_views()

        self.lbl_status.config(text=f"Config: {config_name} - Select a key to view diff")

    def filter_config_keys(self, *args):
        """Filter config keys based on search text"""
        if not hasattr(self, 'all_keys'):
            return

        search_text = self.search_var.get().lower()

        if not search_text:
            filtered_keys = self.all_keys
        else:
            filtered_keys = [k for k in self.all_keys if search_text in str(k).lower()]

        self.list_config_keys.delete(0, tk.END)
        for key in filtered_keys:
            self.list_config_keys.insert(tk.END, key)

        self.lbl_key_count.config(text=f"{len(filtered_keys)} / {len(self.all_keys)} keys")

    def on_select_config_key(self, event):
        """When config key is selected, show diff"""
        selection = self.list_config_keys.curselection()
        if not selection:
            return

        config_key = self.list_config_keys.get(selection[0])
        self.current_config_key = config_key

        composite_key = (self.current_config_name, config_key)

        if composite_key not in self.config_data_index:
            return

        config_data = self.config_data_index[composite_key]

        old_obj = safe_parse(config_data['old'])
        new_obj = safe_parse(config_data['new'])

        self.display_json(old_obj, new_obj)
        self.display_diff(old_obj, new_obj, composite_key)

        self.lbl_status.config(text=f"Showing: {self.current_config_name} | {config_key}")

    def clear_diff_views(self):
        """Clear all diff display areas"""
        self.tree.delete(*self.tree.get_children())
        self.txt_inline_old.delete("1.0", tk.END)
        self.txt_inline_new.delete("1.0", tk.END)
        self.txt_old.delete("1.0", tk.END)
        self.txt_new.delete("1.0", tk.END)

    def refresh_current(self):
        if self.current_config_name and self.current_config_key:
            self.on_select_config_key(None)

    def display_json(self, old, new):
        self.txt_old.delete("1.0", tk.END)
        self.txt_new.delete("1.0", tk.END)

        self.txt_old.insert("1.0", pretty_json(old))
        self.txt_new.insert("1.0", pretty_json(new))

    def display_diff(self, old, new, composite_key):
        self.tree.delete(*self.tree.get_children())

        diff = DeepDiff(old, new, ignore_order=True, verbose_level=2)

        count = 0
        diff_list = []

        # Values changed
        for path, change in diff.get("values_changed", {}).items():
            count += 1
            diff_list.append({
                'path': dd_path_to_key(path),
                'type': 'changed',
                'old': change.get('old_value'),
                'new': change.get('new_value')
            })
            self.tree.insert("", "end", 
                           values=(count, dd_path_to_key(path), "CHANGED", 
                                  str(change.get('old_value'))[:100], 
                                  str(change.get('new_value'))[:100]),
                           tags=('changed',))

        # Type changes
        for path, change in diff.get("type_changes", {}).items():
            count += 1
            diff_list.append({
                'path': dd_path_to_key(path),
                'type': 'changed',
                'old': change.get('old_value'),
                'new': change.get('new_value')
            })
            self.tree.insert("", "end",
                           values=(count, dd_path_to_key(path), "TYPE CHANGE",
                                  str(change.get('old_value'))[:100],
                                  str(change.get('new_value'))[:100]),
                           tags=('changed',))

        # Added
        for path in diff.get("dictionary_item_added", set()):
            count += 1
            diff_list.append({
                'path': dd_path_to_key(path),
                'type': 'added',
                'old': None,
                'new': "ADDED"
            })
            self.tree.insert("", "end",
                           values=(count, dd_path_to_key(path), "ADDED", "", "ADDED"),
                           tags=('added',))

        # Removed
        for path in diff.get("dictionary_item_removed", set()):
            count += 1
            diff_list.append({
                'path': dd_path_to_key(path),
                'type': 'removed',
                'old': "REMOVED",
                'new': None
            })
            self.tree.insert("", "end",
                           values=(count, dd_path_to_key(path), "REMOVED", "REMOVED", ""),
                           tags=('removed',))

        self.config_diffs[composite_key] = diff_list

        self.lbl_status.config(text=f"Found {count} changes for {self.current_config_name} | {self.current_config_key}")

    def on_tree_select(self, event):
        selection = self.tree.selection()
        if not selection:
            return

        item = self.tree.item(selection[0])
        values = item['values']

        if len(values) < 5:
            return

        path = values[1]
        old_val = values[3]
        new_val = values[4]

        # Inline diff
        self.txt_inline_old.delete("1.0", tk.END)
        self.txt_inline_new.delete("1.0", tk.END)

        old_str = str(old_val)
        new_str = str(new_val)

        sm = difflib.SequenceMatcher(a=old_str, b=new_str)

        for op, i1, i2, j1, j2 in sm.get_opcodes():
            if op == 'equal':
                self.txt_inline_old.insert(tk.END, old_str[i1:i2], 'equal')
                self.txt_inline_new.insert(tk.END, new_str[j1:j2], 'equal')
            elif op == 'delete':
                self.txt_inline_old.insert(tk.END, old_str[i1:i2], 'delete')
            elif op == 'insert':
                self.txt_inline_new.insert(tk.END, new_str[j1:j2], 'insert')
            elif op == 'replace':
                self.txt_inline_old.insert(tk.END, old_str[i1:i2], 'delete')
                self.txt_inline_new.insert(tk.END, new_str[j1:j2], 'insert')

        # Highlight ALL changes
        composite_key = (self.current_config_name, self.current_config_key)
        if composite_key in self.config_diffs:
            all_diffs = self.config_diffs[composite_key]
            for diff_item in all_diffs:
                leaf_key = diff_item['path'].split('.')[-1].split('[')[0]
                if leaf_key:
                    self._highlight_line(self.txt_old, leaf_key)
                    self._highlight_line(self.txt_new, leaf_key)

    def _highlight_line(self, widget, key):
        widget.tag_remove('highlight', "1.0", tk.END)

        text = widget.get("1.0", "end-1c")
        if not text.strip() or not key:
            return

        lines = text.split('\n')
        for i, line in enumerate(lines):
            if key in line:
                line_start = f"{i+1}.0"
                line_end = f"{i+1}.end"
                widget.tag_add('highlight', line_start, line_end)

if __name__ == "__main__":
    try:
        app = GeminiUltraProEnhanced()
        app.mainloop()
    except Exception as e:
        print(f"ERROR: {e}")
        logger.error(f"App failed: {e}", exc_info=True)
        input("Press Enter to exit...")
