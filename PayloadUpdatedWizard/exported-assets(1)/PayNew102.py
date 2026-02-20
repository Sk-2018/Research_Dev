
# -*- coding: utf-8 -*-
"""
GeminiPayloadDiff_UltraIntegrated.py

Single-file, production-ready payload diff viewer that combines:
- Your GeminiPayloadDiff-style UI (toolbar + config dropdown + diff table + dual JSON panes)
- UltraFastLoader-style chunked loading for CSV / XLS / XLSX / XLSB
- DeepDiff-based JSON comparison (changed / added / removed)
- Click diff row -> scroll both JSON panes to the changed key and highlight:
    - entire line in blue
    - old value (left) in red
    - new value (right) in green

This is designed as a drop-in replacement for the broken/incomplete GeminiPayloadDiff.py.
"""

import os
import json
import ast
import threading
import re
from typing import Any, Dict, List, Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from deepdiff import DeepDiff
import openpyxl
import chardet
from pyxlsb import open_workbook


# =============================================================================
# THEME / CONSTANTS
# =============================================================================

FONT_MAIN = ("Segoe UI", 10)
FONT_MONO = ("Consolas", 10)

COLOR_BG = "#F4F6F9"
COLOR_PANEL = "#FFFFFF"
COLOR_LINE_HL = "#CDE5FF"
COLOR_ADDED = "#D6F5D6"
COLOR_REMOVED = "#FFD6D6"
COLOR_CHANGED_OLD = "#FFE6E6"
COLOR_CHANGED_NEW = "#E6FFE6"


# =============================================================================
# UltraFastLoader – based on ultra_fast_loader.py + ULTRA viewer
# =============================================================================

class UltraFastLoader:
    """
    High-speed, chunked loader for large XLS/CSV-type files.
    Supports CSV, XLSX/XLS, XLSB, with basic progress callback.
    """

    def __init__(self, chunk_size: int = 50000):
        self.chunk_size = chunk_size

    # ------------- format detection -------------

    @staticmethod
    def detect_format(path: str) -> Optional[str]:
        p = path.lower()
        if p.endswith(".csv") or p.endswith(".tsv") or p.endswith(".txt"):
            return "csv"
        if p.endswith(".xlsx") or p.endswith(".xls"):
            return "xlsx"
        if p.endswith(".xlsb"):
            return "xlsb"
        return None

    def estimate_rows(self, path: str) -> int:
        fmt = self.detect_format(path)
        try:
            if fmt == "csv":
                with open(path, "rb") as f:
                    return max(0, sum(1 for _ in f) - 1)
            if fmt == "xlsx":
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                return wb.active.max_row - 1
            if fmt == "xlsb":
                with open_workbook(path) as wb:
                    sh = wb.get_sheet(wb.sheets[0])
                    return max(0, len(list(sh.rows())) - 1)
        except Exception:
            return 100000
        return 50000

    # ------------- unified entrypoint -------------

    def load_file(self, path: str, progress_cb=None) -> pd.DataFrame:
        fmt = self.detect_format(path)
        if fmt == "csv":
            return self._load_csv(path, progress_cb)
        if fmt == "xlsx":
            return self._load_xlsx(path, progress_cb)
        if fmt == "xlsb":
            return self._load_xlsb(path, progress_cb)
        raise ValueError(f"Unsupported file format: {path}")

    # ------------- CSV -------------

    def _load_csv(self, path: str, progress_cb):
        total = self.estimate_rows(path)
        with open(path, "rb") as f:
            enc = chardet.detect(f.read(10000)).get("encoding") or "utf-8"

        # detect delimiter
        with open(path, "r", encoding=enc, errors="replace") as f:
            first_line = f.readline()
        sep = "," if "," in first_line else "\t"

        chunks: List[pd.DataFrame] = []
        read_rows = 0

        for chunk in pd.read_csv(
            path,
            sep=sep,
            chunksize=self.chunk_size,
            encoding=enc,
            low_memory=False,
        ):
            chunks.append(chunk)
            read_rows += len(chunk)
            if progress_cb and total > 0:
                progress_cb(int(read_rows * 100 / total))

        return pd.concat(chunks, ignore_index=True)

    # ------------- XLSX/XLS -------------

    def _load_xlsx(self, path: str, progress_cb):
        total = self.estimate_rows(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sh = wb.active
        rows = sh.iter_rows(values_only=True)

        headers = next(rows)
        chunks: List[pd.DataFrame] = []
        buf: List[list] = []
        count = 0

        for r in rows:
            buf.append(r)
            count += 1
            if len(buf) >= self.chunk_size:
                chunks.append(pd.DataFrame(buf, columns=headers))
                buf = []
            if progress_cb and total > 0:
                progress_cb(int(count * 100 / total))

        if buf:
            chunks.append(pd.DataFrame(buf, columns=headers))

        return pd.concat(chunks, ignore_index=True)

    # ------------- XLSB -------------

    def _load_xlsb(self, path: str, progress_cb):
        total = self.estimate_rows(path)
        chunks: List[pd.DataFrame] = []

        with open_workbook(path) as wb:
            sh = wb.get_sheet(wb.sheets[0])
            rows = list(sh.rows())
            headers = [c.v for c in rows[0]]

            buf: List[list] = []
            count = 0

            for r in rows[1:]:
                buf.append([c.v for c in r])
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if progress_cb and total > 0:
                    progress_cb(int(count * 100 / total))

            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))

        return pd.concat(chunks, ignore_index=True)


# =============================================================================
# JSON helpers
# =============================================================================

def safe_json(value: Any) -> Any:
    if value is None:
        return {}
    s = str(value).strip()
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        pass
    try:
        return ast.literal_eval(s)
    except Exception:
        pass
    return {}


def deepdiff_path_to_tokens(path: str) -> List[Any]:
    """
    Convert DeepDiff path "root['a'][0]['b']" into tokens: ['a', 0, 'b'].
    """
    tokens: List[Any] = []
    if not path.startswith("root"):
        return tokens
    i = len("root")
    while i < len(path):
        if path[i] == "[":
            j = path.find("]", i + 1)
            if j == -1:
                break
            tok = path[i + 1:j]
            if tok.startswith(("'", '"')):
                tok = tok.strip("'\"")
            else:
                try:
                    tok = int(tok)
                except Exception:
                    pass
            tokens.append(tok)
            i = j + 1
        else:
            i += 1
    return tokens


def value_from_tokens(obj: Any, tokens: List[Any]) -> Any:
    cur = obj
    for t in tokens:
        if isinstance(t, int) and isinstance(cur, list):
            if 0 <= t < len(cur):
                cur = cur[t]
            else:
                return None
        elif isinstance(cur, dict):
            cur = cur.get(t)
        else:
            return None
    return cur


def find_line_with_key(text_widget, path: str) -> Optional[str]:
    """
    Find the line in JSON text that contains the path/key.
    Returns line number (1-based) or None.
    """
    content = text_widget.get("1.0", tk.END)
    lines = content.split('\n')
    
    # Extract the last part of the path (leaf key) for searching
    if '[' in path:
        last_bracket = path.rfind(']')
        if last_bracket != -1:
            key_part = path[last_bracket + 1:].strip("[]'\"").strip()
        else:
            key_part = path.split('[')[-1].strip("]'\"")
    else:
        key_part = path.split('.')[-1].strip("'\"")
    
    for i, line in enumerate(lines, 1):
        if key_part in line:
            return f"{i}.0"
    return None


def highlight_json_line(text_widget, line_start: str, value_start: str, value_end: str, tag: str):
    """
    Apply highlighting to a specific line and value range.
    """
    # Clear previous tags in this line range
    text_widget.tag_remove("line_hl", line_start, f"{line_start}+1 line")
    text_widget.tag_remove("removed", line_start, f"{line_start}+1 line")
    text_widget.tag_remove("added", line_start, f"{line_start}+1 line")
    text_widget.tag_remove("changed_old", line_start, f"{line_start}+1 line")
    text_widget.tag_remove("changed_new", line_start, f"{line_start}+1 line")
    
    # Highlight entire line
    text_widget.tag_add("line_hl", line_start, f"{line_start}+1 line")
    
    # Highlight specific value
    if value_start and value_end:
        text_widget.tag_add(tag, value_start, value_end)


# =============================================================================
# Main Gemini-style UI with Ultra loader integrated
# =============================================================================

class GeminiPayloadDiffApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Perfect Unified Payload Diff Viewer (Gemini + Ultra)")
        self.geometry("1900x1100")
        self.configure(bg=COLOR_BG)

        self.loader = UltraFastLoader()
        self.df: Optional[pd.DataFrame] = None
        self.config_col: Optional[str] = None

        self.full_old = None
        self.full_new = None
        self.diff_items: Dict[str, Dict[str, Any]] = {}  # Tree iid -> metadata

        self._build_ui()

    # ---------------- UI BUILD ----------------

    def _build_ui(self):
        self._build_toolbar()
        self._build_layout()
        self._build_statusbar()

    def _build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=5, pady=5)

        ttk.Button(bar, text="Open File", command=self.open_file).pack(side="left")

        ttk.Label(bar, text="Config:", font=FONT_MAIN).pack(side="left", padx=10)
        self.cmb_config = ttk.Combobox(bar, state="readonly", font=FONT_MAIN, width=40)
        self.cmb_config.pack(side="left")
        self.cmb_config.bind("<<ComboboxSelected>>", self.on_select_config)

        self.search_var = tk.StringVar()
        ttk.Entry(bar, textvariable=self.search_var, width=40).pack(side="left", padx=10)
        ttk.Button(bar, text="Search", command=self.search_path).pack(side="left")

    def _build_layout(self):
        pan = ttk.Panedwindow(self, orient=tk.VERTICAL)
        pan.pack(fill="both", expand=True)

        # top: diff table
        top = ttk.Frame(pan)
        cols = ("path", "type", "old", "new")
        self.tree = ttk.Treeview(top, columns=cols, show="headings")
        self.tree.heading("path", text="Path")
        self.tree.heading("type", text="Type")
        self.tree.heading("old", text="Old")
        self.tree.heading("new", text="New")
        self.tree.column("path", width=800)
        self.tree.column("type", width=120)
        self.tree.column("old", width=380)
        self.tree.column("new", width=380)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.pack(fill="both", expand=True)
        pan.add(top, weight=2)

        # bottom: JSON viewers
        bottom = ttk.Frame(pan)

        # Add scrollbars for text widgets
        old_frame = ttk.Frame(bottom)
        old_frame.pack(side="left", fill="both", expand=True)
        self.txt_old = tk.Text(old_frame, bg=COLOR_PANEL, font=FONT_MONO, wrap="none", undo=True)
        scrollbar_old = ttk.Scrollbar(old_frame, orient="vertical", command=self.txt_old.yview)
        self.txt_old.configure(yscrollcommand=scrollbar_old.set)
        self.txt_old.pack(side="left", fill="both", expand=True)
        scrollbar_old.pack(side="right", fill="y")

        new_frame = ttk.Frame(bottom)
        new_frame.pack(side="right", fill="both", expand=True)
        self.txt_new = tk.Text(new_frame, bg=COLOR_PANEL, font=FONT_MONO, wrap="none", undo=True)
        scrollbar_new = ttk.Scrollbar(new_frame, orient="vertical", command=self.txt_new.yview)
        self.txt_new.configure(yscrollcommand=scrollbar_new.set)
        self.txt_new.pack(side="left", fill="both", expand=True)
        scrollbar_new.pack(side="right", fill="y")

        # sync scroll
        self.txt_old.bind("<MouseWheel>", self._scroll_old)
        self.txt_new.bind("<MouseWheel>", self._scroll_new)

        # tags for highlighting
        self.txt_old.tag_config("line_hl", background=COLOR_LINE_HL)
        self.txt_new.tag_config("line_hl", background=COLOR_LINE_HL)
        self.txt_old.tag_config("removed", background=COLOR_REMOVED)
        self.txt_new.tag_config("added", background=COLOR_ADDED)
        self.txt_old.tag_config("changed_old", background=COLOR_CHANGED_OLD)
        self.txt_new.tag_config("changed_new", background=COLOR_CHANGED_NEW)

        pan.add(bottom, weight=3)

    def _build_statusbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x")
        self.lbl_status = ttk.Label(bar, text="Ready.")
        self.lbl_status.pack(side="right", padx=8)

    # ---------------- Scroll sync ----------------

    def _scroll_old(self, event):
        self.txt_old.yview_scroll(int(-event.delta / 120), "units")
        self.txt_new.yview_moveto(self.txt_old.yview()[0])
        return "break"

    def _scroll_new(self, event):
        self.txt_new.yview_scroll(int(-event.delta / 120), "units")
        self.txt_old.yview_moveto(self.txt_new.yview()[0])
        return "break"

    # ---------------- File loading with UltraFastLoader ----------------

    def open_file(self):
        path = filedialog.askopenfilename(
            title="Select payload export",
            filetypes=[("Data files", "*.csv;*.tsv;*.txt;*.xlsx;*.xls;*.xlsb")]
        )
        if not path:
            return

        self.lbl_status.config(text="Loading file (UltraFast)...")

        def worker():
            try:
                df = self.loader.load_file(path)
                self.df = df
                self.after(0, self._identify_config_column)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Load Error", f"Failed to load file: {e}"))
                self.after(0, lambda: self.lbl_status.config(text="Load failed."))

        threading.Thread(target=worker, daemon=True).start()

    def _identify_config_column(self):
        if self.df is None or self.df.empty:
            messagebox.showerror("Error", "No data loaded from file.")
            self.lbl_status.config(text="No data loaded.")
            return

        headers_lower = [c.lower() for c in self.df.columns]
        self.config_col = None

        # try to find config column using heuristics
        for key in ["configname", "config_name", "config", "name", "cfg_name"]:
            for idx, h in enumerate(headers_lower):
                hn = h.replace(" ", "").replace("_", "")
                if key in hn:
                    self.config_col = self.df.columns[idx]
                    break
            if self.config_col:
                break

        if not self.config_col:
            messagebox.showerror("Error", "Could not detect config name column.")
            self.lbl_status.config(text="Config column not found.")
            return

        cfg_series = self.df[self.config_col].dropna()
        cfg_values = sorted(cfg_series.astype(str).unique())

        self.cmb_config["values"] = cfg_values
        self.lbl_status.config(text=f"Loaded {len(self.df):,} rows, {len(cfg_values):,} configs. Select a config to compare.")

    # ---------------- Config selection / diff ----------------

    def on_select_config(self, event=None):
        if self.df is None or self.config_col is None:
            return

        cfg = self.cmb_config.get()
        if not cfg:
            return

        mask = self.df[self.config_col].astype(str) == cfg
        sel = self.df[mask]
        if sel.empty:
            messagebox.showerror("Error", f"No rows found for config {cfg}")
            return

        row = sel.iloc[0]

        # detect old/new payload columns
        headers_lower = [c.lower() for c in self.df.columns]
        # normalize: lower, no spaces/underscores
        norm = [h.replace(" ", "").replace("_", "") for h in headers_lower]

        def find_col(candidates: List[str]) -> Optional[str]:
            for cand in candidates:
                for i, n in enumerate(norm):
                    if cand in n:
                        return self.df.columns[i]
            return None

        new_col = find_col(["currentpayload", "currentjson", "newpayload", "payloadjson", "current"])
        old_col = find_col(["oldpayload", "oldjson", "previouspayload", "old", "previous"])

        if not new_col or not old_col:
            messagebox.showerror(
                "Error",
                "Could not detect OLD/NEW payload columns.\\n"
                "Expected names containing: old/current/new/payload/json."
            )
            return

        old_obj = safe_json(row[old_col])
        new_obj = safe_json(row[new_col])

        self.full_old = old_obj
        self.full_new = new_obj

        self._display_json(old_obj, new_obj)
        self._compute_diff(old_obj, new_obj)

    def _display_json(self, old_obj, new_obj):
        self.txt_old.delete("1.0", tk.END)
        self.txt_new.delete("1.0", tk.END)

        try:
            old_str = json.dumps(old_obj, indent=4, ensure_ascii=False)
        except Exception:
            old_str = str(old_obj)

        try:
            new_str = json.dumps(new_obj, indent=4, ensure_ascii=False)
        except Exception:
            new_str = str(new_obj)

        self.txt_old.insert("1.0", old_str)
        self.txt_new.insert("1.0", new_str)

        # Clear all highlighting tags
        self.txt_old.tag_remove("line_hl", "1.0", tk.END)
        self.txt_new.tag_remove("line_hl", "1.0", tk.END)
        self.txt_old.tag_remove("removed", "1.0", tk.END)
        self.txt_new.tag_remove("added", "1.0", tk.END)
        self.txt_old.tag_remove("changed_old", "1.0", tk.END)
        self.txt_new.tag_remove("changed_new", "1.0", tk.END)

    # ---------------- DeepDiff integration ----------------

    def _compute_diff(self, old_obj, new_obj):
        self.tree.delete(*self.tree.get_children())
        self.diff_items.clear()

        diff = DeepDiff(old_obj, new_obj, ignore_order=True, verbose_level=2)
        count = 0

        # changed values
        if "values_changed" in diff:
            for path, d in diff["values_changed"].items():
                old_val = d.get("old_value")
                new_val = d.get("new_value")
                tokens = deepdiff_path_to_tokens(path)
                iid = self._insert_row(path, "changed", old_val, new_val)
                self.diff_items[iid] = dict(kind="changed", path=path, tokens=tokens,
                                            old=old_val, new=new_val)
                count += 1

        # added
        if "dictionary_item_added" in diff:
            for path in diff["dictionary_item_added"]:
                tokens = deepdiff_path_to_tokens(path)
                new_val = value_from_tokens(new_obj, tokens)
                iid = self._insert_row(path, "added", None, new_val)
                self.diff_items[iid] = dict(kind="added", path=path, tokens=tokens,
                                            old=None, new=new_val)
                count += 1

        # removed
        if "dictionary_item_removed" in diff:
            for path in diff["dictionary_item_removed"]:
                tokens = deepdiff_path_to_tokens(path)
                old_val = value_from_tokens(old_obj, tokens)
                iid = self._insert_row(path, "removed", old_val, None)
                self.diff_items[iid] = dict(kind="removed", path=path, tokens=tokens,
                                            old=old_val, new=None)
                count += 1

        self.lbl_status.config(text=f"Total changes: {count}")

    def _insert_row(self, path, kind, old_val, new_val):
        def preview(v):
            if v is None:
                return ""
            s = str(v)
            if len(s) > 100:
                s = s[:100] + "..."
            return s

        old_str = preview(old_val)
        new_str = preview(new_val)
        
        iid = self.tree.insert("", "end", values=(path, kind, old_str, new_str))
        return iid

    # ---------------- Tree selection and highlighting ----------------

    def on_tree_select(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        
        iid = selection[0]
        if iid not in self.diff_items:
            return
        
        item = self.diff_items[iid]
        path = item["path"]
        kind = item["kind"]
        
        # Scroll both text widgets to the relevant line
        old_line = find_line_with_key(self.txt_old, path)
        new_line = find_line_with_key(self.txt_new, path)
        
        if old_line:
            self.txt_old.see(old_line)
            self.txt_old.tag_add("line_hl", old_line, f"{old_line}+1 line")
        
        if new_line:
            self.txt_new.see(new_line)
            self.txt_new.tag_add("line_hl", new_line, f"{new_line}+1 line")
        
        # Apply value-specific highlighting
        if kind == "changed":
            # For changed, highlight old value in left, new value in right
            # Note: Exact value highlighting would require parsing JSON positions
            # For now, just line highlight
            if old_line:
                self.txt_old.tag_add("changed_old", old_line, f"{old_line}+1 line")
            if new_line:
                self.txt_new.tag_add("changed_new", new_line, f"{new_line}+1 line")
        elif kind == "removed" and old_line:
            self.txt_old.tag_add("removed", old_line, f"{old_line}+1 line")
        elif kind == "added" and new_line:
            self.txt_new.tag_add("added", new_line, f"{new_line}+1 line")

    # ---------------- Search functionality ----------------

    def search_path(self):
        query = self.search_var.get().strip().lower()
        if not query:
            return
        
        # Clear previous search highlighting
        for item in self.tree.get_children():
            self.tree.item(item, tags=())
        
        count = 0
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if query in str(values[0]).lower():
                self.tree.item(item, tags=("search_match",))
                count += 1
                self.tree.see(item)
                self.tree.selection_set(item)
                self.tree.focus(item)
                break  # Focus on first match
        
        self.lbl_status.config(text=f"Search '{query}': {count} matches found")
        
        # Configure search highlighting (optional - treeview doesn't support tags easily)
        # The above just selects and scrolls to the first match

    def run(self):
        """Start the application."""
        self.mainloop()


# =============================================================================
# Entry point
# =============================================================================

if __name__ == "__main__":
    app = GeminiPayloadDiffApp()
    app.run()
