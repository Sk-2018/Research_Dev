# -*- coding: utf-8 -*-
"""
GeminiPayloadDiff_UltraIntegrated.py

Single-file, production-ready payload diff viewer that combines:
- Your GeminiPayloadDiff-style UI (toolbar + config dropdown + diff table + dual JSON panes)
- UltraFastLoader-style chunked loading for CSV / XLS / XLSX / XLSB
- DeepDiff-based JSON comparison (changed / added / removed)
- Click diff row -> scroll both JSON panes to the changed key and highlight:
    - entire line in blue
    - old value (left) in red
    - new value (right) in green

This is designed as a drop-in replacement for the broken/incomplete GeminiPayloadDiff.py.
"""

import os
import json
import ast
import threading
import difflib
from typing import Any, Dict, List, Optional
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from deepdiff import DeepDiff
import openpyxl
import chardet

try:
    from pyxlsb import open_workbook
    HAS_XLSB = True
except ImportError:
    HAS_XLSB = False


# =============================================================================
# THEME / CONSTANTS
# =============================================================================

FONT_MAIN = ("Segoe UI", 10)
FONT_MONO = ("Consolas", 10)

COLOR_BG = "#F4F6F9"
COLOR_PANEL = "#FFFFFF"
COLOR_LINE_HL = "#CDE5FF"
COLOR_ADDED = "#D6F5D6"
COLOR_REMOVED = "#FFD6D6"


# =============================================================================
# UltraFastLoader – based on ultra_fast_loader.py + ULTRA viewer
# =============================================================================

class UltraFastLoader:
    """
    High-speed, chunked loader for large XLS/CSV-type files.
    Supports CSV, XLSX/XLS, XLSB, with basic progress callback.
    """

    def __init__(self, chunk_size: int = 50000):
        self.chunk_size = chunk_size

    # ------------- format detection -------------

    @staticmethod
    def detect_format(path: str) -> Optional[str]:
        p = path.lower()
        if p.endswith(".csv") or p.endswith(".tsv") or p.endswith(".txt"):
            return "csv"
        if p.endswith(".xlsx") or p.endswith(".xls"):
            return "xlsx"
        if p.endswith(".xlsb"):
            return "xlsb"
        return None

    def estimate_rows(self, path: str) -> int:
        fmt = self.detect_format(path)
        try:
            if fmt == "csv":
                with open(path, "rb") as f:
                    return max(0, sum(1 for _ in f) - 1)
            if fmt == "xlsx":
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                return wb.active.max_row - 1
            if fmt == "xlsb" and HAS_XLSB:
                with open_workbook(path) as wb:
                    sh = wb.get_sheet(wb.sheets[0])
                    return max(0, len(list(sh.rows())) - 1)
        except Exception:
            return 100000
        return 50000

    # ------------- unified entrypoint -------------

    def load_file(self, path: str, progress_cb=None) -> pd.DataFrame:
        fmt = self.detect_format(path)
        if fmt == "csv":
            return self._load_csv(path, progress_cb)
        if fmt == "xlsx":
            return self._load_xlsx(path, progress_cb)
        if fmt == "xlsb":
            return self._load_xlsb(path, progress_cb)
        raise ValueError(f"Unsupported file format: {path}")

    # ------------- CSV -------------

    def _load_csv(self, path: str, progress_cb):
        total = self.estimate_rows(path)
        with open(path, "rb") as f:
            enc = chardet.detect(f.read(10000)).get("encoding") or "utf-8"

        # detect delimiter
        with open(path, "r", encoding=enc, errors="replace") as f:
            first_line = f.readline()
        sep = "," if "," in first_line else "\t"

        chunks: List[pd.DataFrame] = []
        read_rows = 0

        for chunk in pd.read_csv(
            path,
            sep=sep,
            chunksize=self.chunk_size,
            encoding=enc,
            low_memory=False,
            on_bad_lines='skip'
        ):
            chunks.append(chunk)
            read_rows += len(chunk)
            if progress_cb and total > 0:
                progress_cb(int(read_rows * 100 / total))

        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()

    # ------------- XLSX/XLS -------------

    def _load_xlsx(self, path: str, progress_cb):
        total = self.estimate_rows(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sh = wb.active
        rows = sh.iter_rows(values_only=True)

        headers = next(rows)
        chunks: List[pd.DataFrame] = []
        buf: List[list] = []
        count = 0

        for r in rows:
            buf.append(r)
            count += 1
            if len(buf) >= self.chunk_size:
                chunks.append(pd.DataFrame(buf, columns=headers))
                buf = []
            if progress_cb and total > 0:
                progress_cb(int(count * 100 / total))

        if buf:
            chunks.append(pd.DataFrame(buf, columns=headers))

        wb.close()
        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()

    # ------------- XLSB -------------

    def _load_xlsb(self, path: str, progress_cb):
        if not HAS_XLSB:
            raise ImportError("pyxlsb not installed. Install with: pip install pyxlsb")
        
        total = self.estimate_rows(path)
        chunks: List[pd.DataFrame] = []

        with open_workbook(path) as wb:
            sh = wb.get_sheet(wb.sheets[0])
            rows = list(sh.rows())
            headers = [c.v for c in rows[0]]

            buf: List[list] = []
            count = 0

            for r in rows[1:]:
                buf.append([c.v for c in r])
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if progress_cb and total > 0:
                    progress_cb(int(count * 100 / total))

            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))

        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()


# =============================================================================
# JSON helpers
# =============================================================================

def safe_json(value: Any) -> Any:
    if value is None:
        return {}
    s = str(value).strip()
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        pass
    try:
        return ast.literal_eval(s)
    except Exception:
        pass
    return {}


def deepdiff_path_to_tokens(path: str) -> List[Any]:
    """
    Convert DeepDiff path "root['a'][0]['b']" into tokens: ['a', 0, 'b'].
    """
    tokens: List[Any] = []
    if not path.startswith("root"):
        return tokens
    i = len("root")
    while i < len(path):
        if path[i] == "[":
            j = path.find("]", i + 1)
            if j == -1:
                break
            tok = path[i + 1:j]
            if tok.startswith(("'", '"')):
                tok = tok.strip("'\"")
            else:
                try:
                    tok = int(tok)
                except Exception:
                    pass
            tokens.append(tok)
            i = j + 1
        else:
            i += 1
    return tokens


def value_from_tokens(obj: Any, tokens: List[Any]) -> Any:
    cur = obj
    for t in tokens:
        if isinstance(t, int) and isinstance(cur, list):
            if 0 <= t < len(cur):
                cur = cur[t]
            else:
                return None
        elif isinstance(cur, dict):
            cur = cur.get(t)
        else:
            return None
    return cur


# =============================================================================
# Main Gemini-style UI with Ultra loader integrated
# =============================================================================

class GeminiPayloadDiffApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Perfect Unified Payload Diff Viewer (Gemini + Ultra)")
        self.geometry("1900x1100")
        self.configure(bg=COLOR_BG)

        self.loader = UltraFastLoader()
        self.df: Optional[pd.DataFrame] = None
        self.config_col: Optional[str] = None

        self.full_old = None
        self.full_new = None
        self.diff_items: Dict[str, Dict[str, Any]] = {}  # Tree iid -> metadata

        self._build_ui()

    # ---------------- UI BUILD ----------------

    def _build_ui(self):
        self._build_toolbar()
        self._build_layout()
        self._build_statusbar()

    def _build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=5, pady=5)

        ttk.Button(bar, text="📁 Open File", command=self.open_file, width=15).pack(side="left", padx=5)

        ttk.Label(bar, text="Config:", font=FONT_MAIN).pack(side="left", padx=10)
        self.cmb_config = ttk.Combobox(bar, state="readonly", font=FONT_MAIN, width=40)
        self.cmb_config.pack(side="left", padx=5)
        self.cmb_config.bind("<<ComboboxSelected>>", self.on_select_config)

        ttk.Label(bar, text="Search:", font=FONT_MAIN).pack(side="left", padx=10)
        self.search_var = tk.StringVar()
        ttk.Entry(bar, textvariable=self.search_var, width=30).pack(side="left", padx=5)
        ttk.Button(bar, text="🔍 Search", command=self.search_path, width=12).pack(side="left", padx=5)

    def _build_layout(self):
        pan = ttk.Panedwindow(self, orient=tk.VERTICAL)
        pan.pack(fill="both", expand=True, padx=5, pady=5)

        # TOP: diff table
        top_frame = ttk.LabelFrame(pan, text="Changes Summary", padding=5)
        
        tree_container = ttk.Frame(top_frame)
        tree_container.pack(fill="both", expand=True)
        
        cols = ("path", "type", "old", "new")
        self.tree = ttk.Treeview(tree_container, columns=cols, show="headings", height=12)
        self.tree.heading("path", text="JSON Path")
        self.tree.heading("type", text="Type")
        self.tree.heading("old", text="Old Value")
        self.tree.heading("new", text="New Value")
        self.tree.column("path", width=800)
        self.tree.column("type", width=120, anchor="center")
        self.tree.column("old", width=380)
        self.tree.column("new", width=380)
        
        # Color tags
        self.tree.tag_configure('changed', background=COLOR_LINE_HL)
        self.tree.tag_configure('added', background=COLOR_ADDED)
        self.tree.tag_configure('removed', background=COLOR_REMOVED)
        
        tree_scroll = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")
        
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        
        pan.add(top_frame, weight=2)

        # MIDDLE: Inline character diff
        inline_frame = ttk.LabelFrame(pan, text="Selected Item - Character-Level Diff", padding=5)
        
        inline_paned = ttk.PanedWindow(inline_frame, orient=tk.HORIZONTAL)
        inline_paned.pack(fill="both", expand=True)
        
        # Old value
        old_inline = ttk.Frame(inline_paned)
        inline_paned.add(old_inline, weight=1)
        ttk.Label(old_inline, text="Old Value", font=("Segoe UI", 10, "bold"), foreground='#C62828').pack()
        
        self.txt_inline_old = tk.Text(old_inline, height=6, wrap='word', font=FONT_MONO, bg=COLOR_PANEL)
        old_inline_scroll = ttk.Scrollbar(old_inline, command=self.txt_inline_old.yview)
        self.txt_inline_old.configure(yscrollcommand=old_inline_scroll.set)
        self.txt_inline_old.pack(side="left", fill="both", expand=True)
        old_inline_scroll.pack(side="right", fill="y")
        
        # New value
        new_inline = ttk.Frame(inline_paned)
        inline_paned.add(new_inline, weight=1)
        ttk.Label(new_inline, text="New Value", font=("Segoe UI", 10, "bold"), foreground='#2E7D32').pack()
        
        self.txt_inline_new = tk.Text(new_inline, height=6, wrap='word', font=FONT_MONO, bg=COLOR_PANEL)
        new_inline_scroll = ttk.Scrollbar(new_inline, command=self.txt_inline_new.yview)
        self.txt_inline_new.configure(yscrollcommand=new_inline_scroll.set)
        self.txt_inline_new.pack(side="left", fill="both", expand=True)
        new_inline_scroll.pack(side="right", fill="y")
        
        pan.add(inline_frame, weight=1)

        # BOTTOM: JSON viewers with SYNCHRONIZED SCROLLING
        bottom_frame = ttk.LabelFrame(pan, text="Full Payloads (Synchronized Scrolling)", padding=5)
        
        json_paned = ttk.PanedWindow(bottom_frame, orient=tk.HORIZONTAL)
        json_paned.pack(fill="both", expand=True)
        
        # OLD payload
        old_frame = ttk.Frame(json_paned)
        json_paned.add(old_frame, weight=1)
        ttk.Label(old_frame, text="OLD Payload", font=("Segoe UI", 10, "bold"), foreground='#C62828').pack()
        
        old_container = ttk.Frame(old_frame)
        old_container.pack(fill="both", expand=True)
        
        self.txt_old = tk.Text(old_container, bg=COLOR_PANEL, font=FONT_MONO, wrap="none")
        self.old_scroll_y = ttk.Scrollbar(old_container, orient="vertical", command=self._on_old_scroll)
        old_scroll_x = ttk.Scrollbar(old_container, orient="horizontal", command=self.txt_old.xview)
        self.txt_old.configure(yscrollcommand=self._update_old_scrollbar, xscrollcommand=old_scroll_x.set)
        
        self.txt_old.grid(row=0, column=0, sticky="nsew")
        self.old_scroll_y.grid(row=0, column=1, sticky="ns")
        old_scroll_x.grid(row=1, column=0, sticky="ew")
        
        old_container.grid_rowconfigure(0, weight=1)
        old_container.grid_columnconfigure(0, weight=1)
        
        # CURRENT payload
        new_frame = ttk.Frame(json_paned)
        json_paned.add(new_frame, weight=1)
        ttk.Label(new_frame, text="CURRENT Payload", font=("Segoe UI", 10, "bold"), foreground='#2E7D32').pack()
        
        new_container = ttk.Frame(new_frame)
        new_container.pack(fill="both", expand=True)
        
        self.txt_new = tk.Text(new_container, bg=COLOR_PANEL, font=FONT_MONO, wrap="none")
        self.new_scroll_y = ttk.Scrollbar(new_container, orient="vertical", command=self._on_new_scroll)
        new_scroll_x = ttk.Scrollbar(new_container, orient="horizontal", command=self.txt_new.xview)
        self.txt_new.configure(yscrollcommand=self._update_new_scrollbar, xscrollcommand=new_scroll_x.set)
        
        self.txt_new.grid(row=0, column=0, sticky="nsew")
        self.new_scroll_y.grid(row=0, column=1, sticky="ns")
        new_scroll_x.grid(row=1, column=0, sticky="ew")
        
        new_container.grid_rowconfigure(0, weight=1)
        new_container.grid_columnconfigure(0, weight=1)
        
        # Synchronized mousewheel
        self.txt_old.bind("<MouseWheel>", self._on_mousewheel_old)
        self.txt_new.bind("<MouseWheel>", self._on_mousewheel_new)
        self.txt_old.bind("<Button-4>", self._on_mousewheel_old)  # Linux
        self.txt_old.bind("<Button-5>", self._on_mousewheel_old)
        self.txt_new.bind("<Button-4>", self._on_mousewheel_new)
        self.txt_new.bind("<Button-5>", self._on_mousewheel_new)

        # Tags for highlighting
        self.txt_old.tag_config("line_hl", background=COLOR_LINE_HL)
        self.txt_new.tag_config("line_hl", background=COLOR_LINE_HL)
        self.txt_old.tag_config("removed", background=COLOR_REMOVED)
        self.txt_new.tag_config("added", background=COLOR_ADDED)

        pan.add(bottom_frame, weight=3)

    def _build_statusbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=5, pady=5)
        self.lbl_status = ttk.Label(bar, text="Ready - Load a file to begin", anchor="w")
        self.lbl_status.pack(fill="x")

    # ---------------- Synchronized Scroll ----------------

    def _on_old_scroll(self, *args):
        self.txt_old.yview(*args)
        self.txt_new.yview(*args)

    def _on_new_scroll(self, *args):
        self.txt_new.yview(*args)
        self.txt_old.yview(*args)

    def _update_old_scrollbar(self, first, last):
        self.old_scroll_y.set(first, last)
        if self.txt_new.yview() != (float(first), float(last)):
            self.txt_new.yview_moveto(first)

    def _update_new_scrollbar(self, first, last):
        self.new_scroll_y.set(first, last)
        if self.txt_old.yview() != (float(first), float(last)):
            self.txt_old.yview_moveto(first)

    def _on_mousewheel_old(self, event):
        delta = -1 if event.delta < 0 or event.num == 5 else 1
        self.txt_old.yview_scroll(delta, "units")
        self.txt_new.yview_scroll(delta, "units")
        return "break"

    def _on_mousewheel_new(self, event):
        delta = -1 if event.delta < 0 or event.num == 5 else 1
        self.txt_new.yview_scroll(delta, "units")
        self.txt_old.yview_scroll(delta, "units")
        return "break"

    # ---------------- File loading ----------------

    def open_file(self):
        path = filedialog.askopenfilename(
            title="Select payload export",
            filetypes=[
                ("All Supported", "*.csv *.tsv *.txt *.xlsx *.xls *.xlsb"),
                ("CSV files", "*.csv *.tsv *.txt"),
                ("Excel files", "*.xlsx *.xls *.xlsb"),
                ("All files", "*.*")
            ]
        )
        if not path:
            return

        self.lbl_status.config(text="Loading file (UltraFast)...")
        self.update_idletasks()

        def worker():
            try:
                df = self.loader.load_file(path)
                self.df = df
                self.after(0, self._identify_config_column)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Load Error", f"Failed to load:\n{str(e)}"))
                self.after(0, lambda: self.lbl_status.config(text="Error loading file"))

        threading.Thread(target=worker, daemon=True).start()

    def _identify_config_column(self):
        if self.df is None or self.df.empty:
            messagebox.showerror("Error", "No data loaded from file.")
            self.lbl_status.config(text="No data loaded.")
            return

        headers_lower = [c.lower() for c in self.df.columns]
        self.config_col = None

        # Find config column
        for key in ["configname", "config_name", "config", "name", "cfg_name", "cfgname"]:
            for idx, h in enumerate(headers_lower):
                hn = h.replace(" ", "").replace("_", "")
                if key in hn:
                    self.config_col = self.df.columns[idx]
                    break
            if self.config_col:
                break

        if not self.config_col:
            messagebox.showerror("Error", "Could not detect config name column.")
            self.lbl_status.config(text="Config column not found.")
            return

        cfg_series = self.df[self.config_col].dropna()
        cfg_values = sorted(cfg_series.astype(str).unique())

        self.cmb_config["values"] = cfg_values
        self.lbl_status.config(text=f"✓ Loaded {len(self.df):,} rows, {len(cfg_values):,} configs. Select a config to compare.")

    # ---------------- Config selection ----------------

    def on_select_config(self, event=None):
        if self.df is None or self.config_col is None:
            return

        cfg = self.cmb_config.get()
        if not cfg:
            return

        mask = self.df[self.config_col].astype(str) == cfg
        sel = self.df[mask]
        if sel.empty:
            messagebox.showerror("Error", f"No rows found for config {cfg}")
            return

        row = sel.iloc[0]

        # Detect old/new columns
        headers_lower = [c.lower() for c in self.df.columns]
        norm = [h.replace(" ", "").replace("_", "") for h in headers_lower]

        def find_col(candidates: List[str]) -> Optional[str]:
            for cand in candidates:
                for i, n in enumerate(norm):
                    if cand in n:
                        return self.df.columns[i]
            return None

        new_col = find_col(["currentpayload", "currentjson", "newpayload", "payloadjson", "current", "payload"])
        old_col = find_col(["oldpayload", "oldjson", "previouspayload", "prevpayload", "old", "previous", "prev"])

        if not new_col or not old_col:
            messagebox.showerror(
                "Error",
                "Could not detect OLD/NEW payload columns.\n"
                "Expected names containing: old/current/new/payload/json/previous."
            )
            return

        old_obj = safe_json(row[old_col])
        new_obj = safe_json(row[new_col])

        self.full_old = old_obj
        self.full_new = new_obj

        self._display_json(old_obj, new_obj)
        self._compute_diff(old_obj, new_obj)

    def _display_json(self, old_obj, new_obj):
        self.txt_old.delete("1.0", tk.END)
        self.txt_new.delete("1.0", tk.END)

        try:
            old_str = json.dumps(old_obj, indent=4, ensure_ascii=False)
        except Exception:
            old_str = str(old_obj)

        try:
            new_str = json.dumps(new_obj, indent=4, ensure_ascii=False)
        except Exception:
            new_str = str(new_obj)

        self.txt_old.insert("1.0", old_str)
        self.txt_new.insert("1.0", new_str)

    # ---------------- DeepDiff ----------------

    def _compute_diff(self, old_obj, new_obj):
        self.tree.delete(*self.tree.get_children())
        self.diff_items.clear()

        diff = DeepDiff(old_obj, new_obj, ignore_order=True, verbose_level=2)
        count = 0

        # Changed values
        if "values_changed" in diff:
            for path, d in diff["values_changed"].items():
                old_val = d.get("old_value")
                new_val = d.get("new_value")
                tokens = deepdiff_path_to_tokens(path)
                iid = self._insert_row(path, "changed", old_val, new_val)
                self.diff_items[iid] = dict(kind="changed", path=path, tokens=tokens,
                                            old=old_val, new=new_val)
                count += 1

        # Added
        if "dictionary_item_added" in diff:
            for path in diff["dictionary_item_added"]:
                tokens = deepdiff_path_to_tokens(path)
                new_val = value_from_tokens(new_obj, tokens)
                iid = self._insert_row(path, "added", None, new_val)
                self.diff_items[iid] = dict(kind="added", path=path, tokens=tokens,
                                            old=None, new=new_val)
                count += 1

        # Removed
        if "dictionary_item_removed" in diff:
            for path in diff["dictionary_item_removed"]:
                tokens = deepdiff_path_to_tokens(path)
                old_val = value_from_tokens(old_obj, tokens)
                iid = self._insert_row(path, "removed", old_val, None)
                self.diff_items[iid] = dict(kind="removed", path=path, tokens=tokens,
                                            old=old_val, new=None)
                count += 1

        self.lbl_status.config(text=f"✓ Total changes: {count}")

    def _insert_row(self, path, kind, old_val, new_val):
        def preview(v):
            if v is None:
                return ""
            try:
                s = json.dumps(v, ensure_ascii=False)
            except Exception:
                s = str(v)
            return s if len(s) <= 120 else s[:117] + "..."

        old_str = preview(old_val)
        new_str = preview(new_val)

        iid = self.tree.insert(
            "",
            "end",
            values=(path, kind.upper(), old_str, new_str),
            tags=(kind,)
        )
        return iid

    # ---------------- Tree selection ----------------

    def on_tree_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        meta = self.diff_items.get(iid)
        if not meta:
            return

        old_val = meta.get("old")
        new_val = meta.get("new")

        # Inline diff
        self._show_inline_diff(old_val, new_val)

        # Full JSON highlighting
        if not self.full_old or not self.full_new:
            return

        tokens = meta["tokens"]
        key_name = str(tokens[-1]) if tokens else None

        self._highlight_json_lines(key_name, old_val, new_val)

    def _show_inline_diff(self, old_val, new_val):
        """Character-level diff."""
        self.txt_inline_old.delete("1.0", tk.END)
        self.txt_inline_new.delete("1.0", tk.END)

        old_str = "" if old_val is None else str(old_val)
        new_str = "" if new_val is None else str(new_val)

        sm = difflib.SequenceMatcher(a=old_str, b=new_str)

        self.txt_inline_old.tag_configure("equal", background=COLOR_PANEL)
        self.txt_inline_old.tag_configure("del", background=COLOR_REMOVED)
        self.txt_inline_new.tag_configure("equal", background=COLOR_PANEL)
        self.txt_inline_new.tag_configure("add", background=COLOR_ADDED)

        for op, i1, i2, j1, j2 in sm.get_opcodes():
            a_chunk = old_str[i1:i2]
            b_chunk = new_str[j1:j2]

            if op == "equal":
                self.txt_inline_old.insert(tk.END, a_chunk, "equal")
                self.txt_inline_new.insert(tk.END, b_chunk, "equal")
            elif op == "delete":
                self.txt_inline_old.insert(tk.END, a_chunk, "del")
            elif op == "insert":
                self.txt_inline_new.insert(tk.END, b_chunk, "add")
            elif op == "replace":
                self.txt_inline_old.insert(tk.END, a_chunk, "del")
                self.txt_inline_new.insert(tk.END, b_chunk, "add")

    def _highlight_json_lines(self, key_name: Optional[str], old_val, new_val):
        """Highlight lines in full JSON."""
        # Clear
        self.txt_old.tag_remove("line_hl", "1.0", tk.END)
        self.txt_old.tag_remove("removed", "1.0", tk.END)
        self.txt_new.tag_remove("line_hl", "1.0", tk.END)
        self.txt_new.tag_remove("added", "1.0", tk.END)

        if key_name is None:
            return

        try:
            old_text = self.txt_old.get("1.0", "end-1c")
            new_text = self.txt_new.get("1.0", "end-1c")
        except tk.TclError:
            return

        old_lines = old_text.split("\n")
        new_lines = new_text.split("\n")

        def val_str(v):
            if v is None:
                return ""
            try:
                return json.dumps(v, ensure_ascii=False)
            except Exception:
                return str(v)

        old_value_str = val_str(old_val)
        new_value_str = val_str(new_val)

        def highlight_widget(widget, lines, value_str, tag_line, tag_val):
            for idx, line in enumerate(lines):
                if key_name in line and (not value_str or value_str in line):
                    line_start = f"{idx + 1}.0"
                    line_end = f"{idx + 1}.end"
                    widget.tag_add(tag_line, line_start, line_end)
                    if value_str:
                        col = line.find(value_str)
                        if col >= 0:
                            val_start = f"{idx + 1}.{col}"
                            val_end = f"{idx + 1}.{col + len(value_str)}"
                            widget.tag_add(tag_val, val_start, val_end)
                    widget.see(line_start)

        highlight_widget(self.txt_old, old_lines, old_value_str, "line_hl", "removed")
        highlight_widget(self.txt_new, new_lines, new_value_str, "line_hl", "added")

    # ---------------- Search ----------------

    def search_path(self):
        """Search by path."""
        term = self.search_var.get().strip()
        if not term:
            return

        for iid in self.tree.get_children(""):
            vals = self.tree.item(iid, "values")
            if vals and term.lower() in str(vals[0]).lower():
                self.tree.selection_set(iid)
                self.tree.see(iid)
                self.on_tree_select()
                return

        messagebox.showinfo("Search", f"No path containing '{term}' found.")


# =============================================================================
# Entry point
# =============================================================================

if __name__ == "__main__":
    app = GeminiPayloadDiffApp()
    app.mainloop()
