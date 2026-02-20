"""
PerfectPayloadDiff_GeminiUltra_Production.py

Single-file payload diff viewer:
- UltraFastLoader (CSV/XLS/XLSX/XLSB)
- DeepDiff JSON diff (changed / added / removed)
- Diff table + dual JSON panes
- Click diff row -> scroll + highlight line (blue) + values (red/green)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json, ast, threading
import pandas as pd
from deepdiff import DeepDiff
import openpyxl
import chardet
from pyxlsb import open_workbook


# ============================= THEME ====================================

FONT_MAIN = ("Segoe UI", 10)
FONT_MONO = ("Consolas", 10)

COLOR_BG = "#F4F6F9"
COLOR_PANEL = "#FFFFFF"
COLOR_LINE_HL = "#CDE5FF"
COLOR_ADDED = "#D6F5D6"
COLOR_REMOVED = "#FFD6D6"


# ========================== UltraFastLoader ==============================

class UltraFastLoader:
    def __init__(self, chunk_size=50000):
        self.chunk_size = chunk_size

    def detect_format(self, path: str):
        p = path.lower()
        if p.endswith(".csv"):
            return "csv"
        if p.endswith(".xlsx") or p.endswith(".xls"):
            return "xlsx"
        if p.endswith(".xlsb"):
            return "xlsb"
        return None

    def estimate_rows(self, path: str) -> int:
        fmt = self.detect_format(path)
        try:
            if fmt == "csv":
                with open(path, "rb") as f:
                    return max(0, sum(1 for _ in f) - 1)
            if fmt == "xlsx":
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                return wb.active.max_row - 1
            if fmt == "xlsb":
                with open_workbook(path) as wb:
                    sh = wb.get_sheet(wb.sheets[0])
                    return max(0, len(list(sh.rows())) - 1)
        except Exception:
            return 100000
        return 50000

    def load_file(self, path, cb=None) -> pd.DataFrame:
        fmt = self.detect_format(path)
        if fmt == "csv":
            return self._load_csv(path, cb)
        if fmt == "xlsx":
            return self._load_xlsx(path, cb)
        if fmt == "xlsb":
            return self._load_xlsb(path, cb)
        raise ValueError("Unsupported format")

    def _load_csv(self, path, cb):
        total = self.estimate_rows(path)
        with open(path, "rb") as f:
            enc = chardet.detect(f.read(10000)).get("encoding") or "utf-8"

        with open(path, "r", encoding=enc, errors="replace") as f:
            first = f.readline()
        sep = "," if "," in first else "\t"

        chunks = []
        rows = 0
        for chunk in pd.read_csv(
            path,
            sep=sep,
            chunksize=self.chunk_size,
            encoding=enc,
            low_memory=False,
        ):
            chunks.append(chunk)
            rows += len(chunk)
            if cb and total > 0:
                cb(int(rows * 100 / total))
        return pd.concat(chunks, ignore_index=True)

    def _load_xlsx(self, path, cb):
        total = self.estimate_rows(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sh = wb.active
        rows = sh.iter_rows(values_only=True)
        headers = next(rows)

        chunks, buf, count = [], [], 0
        for r in rows:
            buf.append(r)
            count += 1
            if len(buf) >= self.chunk_size:
                chunks.append(pd.DataFrame(buf, columns=headers))
                buf = []
            if cb and total > 0:
                cb(int(count * 100 / total))
        if buf:
            chunks.append(pd.DataFrame(buf, columns=headers))
        return pd.concat(chunks, ignore_index=True)

    def _load_xlsb(self, path, cb):
        total = self.estimate_rows(path)
        chunks = []
        with open_workbook(path) as wb:
            sh = wb.get_sheet(wb.sheets[0])
            rows = list(sh.rows())
            headers = [c.v for c in rows[0]]

            buf, count = [], 0
            for r in rows[1:]:
                buf.append([c.v for c in r])
                count += 1
                if len(buf) >= self.chunk_size:
                    chunks.append(pd.DataFrame(buf, columns=headers))
                    buf = []
                if cb and total > 0:
                    cb(int(count * 100 / total))
            if buf:
                chunks.append(pd.DataFrame(buf, columns=headers))
        return pd.concat(chunks, ignore_index=True)


# =========================== JSON helpers ================================

def safe_json(text):
    if text is None:
        return {}
    s = str(text).strip()
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        pass
    try:
        return ast.literal_eval(s)
    except Exception:
        pass
    return {}


def deepdiff_path_to_tokens(path: str):
    """
    Convert DeepDiff path "root['a'][0]['b']" to tokens: ['a', 0, 'b']
    """
    tokens = []
    if not path.startswith("root"):
        return tokens
    i = len("root")
    while i < len(path):
        if path[i] == "[":
            j = path.find("]", i + 1)
            if j == -1:
                break
            tok = path[i + 1:j]
            if tok.startswith(("'", '"')):
                tok = tok.strip("'\"")
            else:
                try:
                    tok = int(tok)
                except Exception:
                    pass
            tokens.append(tok)
            i = j + 1
        else:
            i += 1
    return tokens


def value_from_tokens(obj, tokens):
    cur = obj
    for t in tokens:
        if isinstance(t, int) and isinstance(cur, list):
            if 0 <= t < len(cur):
                cur = cur[t]
            else:
                return None
        elif isinstance(cur, dict):
            cur = cur.get(t)
        else:
            return None
    return cur


# ========================== Main Application ============================

class PayloadDiffApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Perfect Payload Diff Viewer (Gemini Pro + Ultra)")
        self.geometry("1900x1100")
        self.configure(bg=COLOR_BG)

        self.loader = UltraFastLoader()
        self.df = None
        self.config_col = None

        self.full_old = None
        self.full_new = None
        self.diff_items = {}  # tree iid -> meta

        self._build_ui()

    # ---------------------- UI ---------------------------------

    def _build_ui(self):
        self._build_toolbar()
        self._build_layout()
        self._build_statusbar()

    def _build_toolbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=5, pady=5)

        ttk.Button(bar, text="Open File", command=self.open_file).pack(side="left")

        ttk.Label(bar, text="Config:", font=FONT_MAIN).pack(side="left", padx=10)
        self.cmb_config = ttk.Combobox(bar, state="readonly", font=FONT_MAIN)
        self.cmb_config.pack(side="left")
        self.cmb_config.bind("<<ComboboxSelected>>", self.on_select_config)

        self.search_var = tk.StringVar()
        ttk.Entry(bar, textvariable=self.search_var, width=40).pack(side="left", padx=10)
        ttk.Button(bar, text="Search", command=self.search_path).pack(side="left")

    def _build_layout(self):
        pan = ttk.Panedwindow(self, orient=tk.VERTICAL)
        pan.pack(fill="both", expand=True)

        # upper diff table
        top = ttk.Frame(pan)
        cols = ("path", "type", "old", "new")
        self.tree = ttk.Treeview(top, columns=cols, show="headings")
        self.tree.heading("path", text="Path")
        self.tree.heading("type", text="Type")
        self.tree.heading("old", text="Old")
        self.tree.heading("new", text="New")
        self.tree.column("path", width=800)
        self.tree.column("type", width=100)
        self.tree.column("old", width=380)
        self.tree.column("new", width=380)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.pack(fill="both", expand=True)
        pan.add(top, weight=2)

        # lower JSON panes
        bottom = ttk.Frame(pan)

        self.txt_old = tk.Text(bottom, bg=COLOR_PANEL, font=FONT_MONO, wrap="none")
        self.txt_new = tk.Text(bottom, bg=COLOR_PANEL, font=FONT_MONO, wrap="none")

        self.txt_old.pack(side="left", fill="both", expand=True)
        self.txt_new.pack(side="right", fill="both", expand=True)

        # sync scroll
        self.txt_old.bind("<MouseWheel>", self._scroll_old)
        self.txt_new.bind("<MouseWheel>", self._scroll_new)

        # tags
        self.txt_old.tag_config("line_hl", background=COLOR_LINE_HL)
        self.txt_new.tag_config("line_hl", background=COLOR_LINE_HL)
        self.txt_old.tag_config("removed", background=COLOR_REMOVED)
        self.txt_new.tag_config("added", background=COLOR_ADDED)

        pan.add(bottom, weight=3)

    def _build_statusbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x")
        self.lbl_status = ttk.Label(bar, text="Ready.")
        self.lbl_status.pack(side="right", padx=8)

    # ------------------- Scroll sync ---------------------------

    def _scroll_old(self, event):
        self.txt_old.yview_scroll(int(-event.delta / 120), "units")
        self.txt_new.yview_moveto(self.txt_old.yview()[0])
        return "break"

    def _scroll_new(self, event):
        self.txt_new.yview_scroll(int(-event.delta / 120), "units")
        self.txt_old.yview_moveto(self.txt_new.yview()[0])
        return "break"

    # ------------------- File loading --------------------------

    def open_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Data files", "*.csv;*.xlsx;*.xls;*.xlsb")]
        )
        if not path:
            return

        self.lbl_status.config(text="Loading file...")

        def worker():
            df = self.loader.load_file(path)
            self.df = df
            self.after(0, self._identify_config_column)

        threading.Thread(target=worker, daemon=True).start()

    def _identify_config_column(self):
        headers_lower = [c.lower() for c in self.df.columns]
        self.config_col = None
        for key in ["config", "config_name", "cfg", "name"]:
            for h in headers_lower:
                if key == h or key in h:
                    self.config_col = self.df.columns[headers_lower.index(h)]
                    break
            if self.config_col:
                break

        if not self.config_col:
            messagebox.showerror("Error", "Config column not found.")
            self.lbl_status.config(text="Config column not found.")
            return

        cfg_series = self.df[self.config_col].dropna()
        cfg_values = sorted(cfg_series.astype(str).unique())
        self.cmb_config["values"] = cfg_values
        self.lbl_status.config(text="File loaded. Select a config.")

    # ------------------- Config selection ----------------------

    def on_select_config(self, event=None):
        if self.df is None or self.config_col is None:
            return
        cfg = self.cmb_config.get()
        if not cfg:
            return

        mask = self.df[self.config_col].astype(str) == cfg
        sel = self.df[mask]
        if sel.empty:
            messagebox.showerror("Error", f"No rows for config {cfg}")
            return

        row = sel.iloc[0]

        headers_lower = [c.lower() for c in self.df.columns]
        new_col = old_col = None

        for key in ["current", "new", "payload_json"]:
            for h in headers_lower:
                if key in h:
                    new_col = self.df.columns[headers_lower.index(h)]
                    break
            if new_col:
                break

        for key in ["old", "previous", "prev_payload"]:
            for h in headers_lower:
                if key in h:
                    old_col = self.df.columns[headers_lower.index(h)]
                    break
            if old_col:
                break

        if not new_col or not old_col:
            messagebox.showerror("Error", "Could not detect old/new payload columns.")
            return

        old_obj = safe_json(row[old_col])
        new_obj = safe_json(row[new_col])
        self.full_old = old_obj
        self.full_new = new_obj

        self._display_json(old_obj, new_obj)
        self._compute_diff(old_obj, new_obj)

    # ------------------- JSON display --------------------------

    def _display_json(self, old_obj, new_obj):
        self.txt_old.delete("1.0", tk.END)
        self.txt_new.delete("1.0", tk.END)

        try:
            old_str = json.dumps(old_obj, indent=4, ensure_ascii=False)
        except Exception:
            old_str = str(old_obj)

        try:
            new_str = json.dumps(new_obj, indent=4, ensure_ascii=False)
        except Exception:
            new_str = str(new_obj)

        self.txt_old.insert("1.0", old_str)
        self.txt_new.insert("1.0", new_str)

    # ------------------- Diff calculation ----------------------

    def _compute_diff(self, old_obj, new_obj):
        self.tree.delete(*self.tree.get_children())
        self.diff_items.clear()

        diff = DeepDiff(old_obj, new_obj, ignore_order=True, verbose_level=2)
        count = 0

        # changed values
        if "values_changed" in diff:
            for path, d in diff["values_changed"].items():
                old_val = d.get("old_value")
                new_val = d.get("new_value")
                tokens = deepdiff_path_to_tokens(path)
                iid = self._insert_row(path, "changed", old_val, new_val)
                self.diff_items[iid] = dict(
                    path=path, kind="changed", tokens=tokens,
                    old=old_val, new=new_val
                )
                count += 1

        # added
        if "dictionary_item_added" in diff:
            for path in diff["dictionary_item_added"]:
                tokens = deepdiff_path_to_tokens(path)
                new_val = value_from_tokens(new_obj, tokens)
                iid = self._insert_row(path, "added", None, new_val)
                self.diff_items[iid] = dict(
                    path=path, kind="added", tokens=tokens,
                    old=None, new=new_val
                )
                count += 1

        # removed
        if "dictionary_item_removed" in diff:
            for path in diff["dictionary_item_removed"]:
                tokens = deepdiff_path_to_tokens(path)
                old_val = value_from_tokens(old_obj, tokens)
                iid = self._insert_row(path, "removed", old_val, None)
                self.diff_items[iid] = dict(
                    path=path, kind="removed", tokens=tokens,
                    old=old_val, new=None
                )
                count += 1

        self.lbl_status.config(text=f"Total changes: {count}")

    def _insert_row(self, path, kind, old_val, new_val):
        def preview(v):
            if v is None:
                return ""
            s = json.dumps(v, ensure_ascii=False)
            return s if len(s) <= 60 else s[:57] + "..."
        return self.tree.insert(
            "", "end",
            values=(path, kind, preview(old_val), preview(new_val))
        )

    # ------------------- Search -------------------------------

    def search_path(self):
        text = self.search_var.get().strip()
        if not text:
            return
        for iid in self.tree.get_children():
            path = self.tree.item(iid)["values"][0]
            if text.lower() in str(path).lower():
                self.tree.selection_set(iid)
                self.tree.see(iid)
                self.on_tree_select()
                return

    # ------------------- Tree selection -----------------------

    def on_tree_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        meta = self.diff_items.get(iid)
        if not meta:
            return
        self._highlight_change(meta)

    # ------------------- Highlight helpers --------------------

    def _clear_highlights(self):
        for w in (self.txt_old, self.txt_new):
            w.tag_remove("line_hl", "1.0", tk.END)
            w.tag_remove("removed", "1.0", tk.END)
            w.tag_remove("added", "1.0", tk.END)

    def _highlight_key_line(self, widget: tk.Text, key: str):
        """
        Highlight line containing "key" in blue.
        """
        if not key:
            return
        search_pat = f'"{key}"'
        idx = widget.search(search_pat, "1.0", stopindex="end")
        if not idx:
            return
        line = idx.split(".")[0]
        widget.tag_add("line_hl", f"{line}.0", f"{line}.0 lineend")
        widget.see(f"{line}.0")

    def _highlight_value(self, widget: tk.Text, value, tag_name: str):
        if value is None:
            return
        s = json.dumps(value, ensure_ascii=False)
        idx = widget.search(s, "1.0", stopindex="end")
        if not idx:
            return
        line, col = idx.split(".")
        start = idx
        end = f"{line}.{int(col) + len(s)}"
        widget.tag_add(tag_name, start, end)

    def _highlight_change(self, meta):
        self._clear_highlights()

        tokens = meta.get("tokens") or []
        key = None
        if tokens:
            last = tokens[-1]
            if isinstance(last, str):
                key = last

        kind = meta["kind"]
        old_val = meta.get("old")
        new_val = meta.get("new")

        # First: highlight key line(s) by key
        if key:
            if kind in ("changed", "removed"):
                self._highlight_key_line(self.txt_old, key)
            if kind in ("changed", "added"):
                self._highlight_key_line(self.txt_new, key)

        # Then: highlight values
        if kind in ("changed", "removed") and old_val is not None:
            self._highlight_value(self.txt_old, old_val, "removed")
        if kind in ("changed", "added") and new_val is not None:
            self._highlight_value(self.txt_new, new_val, "added")


# ============================= MAIN =====================================

if __name__ == "__main__":
    app = PayloadDiffApp()
    app.mainloop()
