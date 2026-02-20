
"""
PerfectUnifiedPayloadDiff_GeminiUltra.py
Unified Payload Diff Viewer
- UltraFastLoader (CSV/XLSX/XLSB)
- Gemini-like UI
- Accurate DeepDiff JSON comparison
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json, ast, threading
import pandas as pd
from deepdiff import DeepDiff
import openpyxl
import chardet
from pyxlsb import open_workbook


# ============================================================
# UltraFastLoader
# ============================================================
class UltraFastLoader:
    def __init__(self, chunk_size=50000):
        self.chunk_size = chunk_size

    def detect_format(self, file):
        f = file.lower()
        if f.endswith(".csv"): return "csv"
        if f.endswith(".xlsx") or f.endswith(".xls"): return "xlsx"
        if f.endswith(".xlsb"): return "xlsb"
        return None

    def estimate_rows(self, file):
        fmt = self.detect_format(file)
        if fmt == "csv":
            with open(file, "rb") as f:
                return sum(1 for _ in f) - 1
        if fmt == "xlsx":
            try:
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                return wb.active.max_row - 1
            except:
                return 100000
        if fmt == "xlsb":
            try:
                with open_workbook(file) as wb:
                    sh = wb.get_sheet(wb.sheets[0])
                    return len(list(sh.rows())) - 1
            except:
                return 100000
        return 50000

    def load_csv(self, file, cb):
        total = self.estimate_rows(file)
        with open(file, 'rb') as f:
            enc = chardet.detect(f.read(10000))["encoding"] or "utf-8"
        # detect separator
        with open(file, "r", encoding=enc, errors="replace") as f:
            first_line = f.readline()
        sep = "," if "," in first_line else "\t"

        chunks = []
        read_rows = 0
        for chunk in pd.read_csv(file, sep=sep, chunksize=self.chunk_size, encoding=enc, low_memory=False):
            chunks.append(chunk)
            read_rows += len(chunk)
            if cb:
                cb(int(read_rows * 100 / total))
        return pd.concat(chunks, ignore_index=True)

    def load_xlsx(self, file, cb):
        total = self.estimate_rows(file)
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        sh = wb.active
        rows = sh.iter_rows(values_only=True)

        headers = next(rows)
        data = []
        chunk = []
        count = 0

        for r in rows:
            chunk.append(r)
            count += 1
            if len(chunk) >= self.chunk_size:
                data.append(pd.DataFrame(chunk, columns=headers))
                chunk = []
            if cb:
                cb(int(count * 100 / total))

        if chunk:
            data.append(pd.DataFrame(chunk, columns=headers))
        return pd.concat(data, ignore_index=True)

    def load_xlsb(self, file, cb):
        total = self.estimate_rows(file)
        data = []
        with open_workbook(file) as wb:
            sh = wb.get_sheet(wb.sheets[0])
            rows = list(sh.rows())
            headers = [c.v for c in rows[0]]
            chunk = []
            count = 0

            for r in rows[1:]:
                chunk.append([c.v for c in r])
                count += 1
                if len(chunk) >= self.chunk_size:
                    data.append(pd.DataFrame(chunk, columns=headers))
                    chunk = []
                if cb:
                    cb(int(count * 100 / total))
            if chunk:
                data.append(pd.DataFrame(chunk, columns=headers))

        return pd.concat(data, ignore_index=True)

    def load_file(self, file, cb=None):
        fmt = self.detect_format(file)
        if fmt == "csv": return self.load_csv(file, cb)
        if fmt == "xlsx": return self.load_xlsx(file, cb)
        if fmt == "xlsb": return self.load_xlsb(file, cb)
        raise ValueError("Unsupported format")


# ============================================================
# Helper JSON parser
# ============================================================
def safe_parse(text):
    if text is None: return {}
    s = str(text).strip()
    if not s: return {}
    try: return json.loads(s)
    except: pass
    try: return ast.literal_eval(s)
    except: pass
    return {}


# ============================================================
# Unified Gemini-Style UI
# ============================================================
class GeminiUltraApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Perfect Unified Payload Diff (Gemini + Ultra)")
        self.geometry("1600x900")

        self.df = None
        self.loader = UltraFastLoader()

        self.create_ui()

    # ------------------------------
    # UI Setup
    # ------------------------------
    def create_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x")

        ttk.Button(top, text="Open File", command=self.open_file).pack(side="left")

        ttk.Label(top, text="Config:").pack(side="left", padx=5)
        self.cmb = ttk.Combobox(top, state="readonly")
        self.cmb.pack(side="left")
        self.cmb.bind("<<ComboboxSelected>>", self.select_config)

        self.status = ttk.Label(top, text="", foreground="blue")
        self.status.pack(side="right")

        # Main splitter
        pan = ttk.Panedwindow(self, orient=tk.VERTICAL)
        pan.pack(fill="both", expand=True)

        # Upper diff summary table
        frame_top = ttk.Frame(pan)
        self.tree = ttk.Treeview(frame_top, columns=("path","type","old","new"), show="headings")
        for c in ("path","type","old","new"):
            self.tree.heading(c, text=c.capitalize())
        self.tree.pack(fill="both", expand=True)
        pan.add(frame_top, weight=1)

        # Lower JSON viewer
        frame_bottom = ttk.Frame(pan)
        self.txt_old = tk.Text(frame_bottom, wrap="none", bg="#f7f7f7")
        self.txt_new = tk.Text(frame_bottom, wrap="none", bg="#f0f0f0")
        self.txt_old.pack(side="left", fill="both", expand=True)
        self.txt_new.pack(side="left", fill="both", expand=True)

        # Sync scroll
        self.txt_old.bind("<MouseWheel>", self.sync_scroll_old)
        self.txt_new.bind("<MouseWheel>", self.sync_scroll_new)

        pan.add(frame_bottom, weight=1)

    # ------------------------------
    # Scroll sync
    # ------------------------------
    def sync_scroll_old(self, e):
        self.txt_new.yview_moveto(self.txt_old.yview()[0])
        return "break"

    def sync_scroll_new(self, e):
        self.txt_old.yview_moveto(self.txt_new.yview()[0])
        return "break"

    # ------------------------------
    # Load file
    # ------------------------------
    def open_file(self):
        file = filedialog.askopenfilename(
            filetypes=[("Data files","*.csv;*.xlsx;*.xlsb;*.xls")]
        )
        if not file: return

        self.status.config(text="Loading...")

        def load():
            df = self.loader.load_file(file)
            self.df = df
            self.after(0, self.populate_configs)

        threading.Thread(target=load).start()

    def populate_configs(self):
        headers = [h.lower() for h in self.df.columns]
        config_col = None
        for candidate in ["config","config_name","cfg","name"]:
            for h in headers:
                if candidate in h:
                    config_col = self.df.columns[headers.index(h)]
                    break
            if config_col: break

        if not config_col:
            messagebox.showerror("Error","Config column not found.")
            return

        self.config_col = config_col
        configs = sorted(self.df[config_col].dropna().unique())
        self.cmb["values"] = configs
        self.status.config(text="Select config")

    # ------------------------------
    # Config selection
    # ------------------------------
    def select_config(self, e):
        cfg = self.cmb.get()
        if not cfg: return
        self.status.config(text="Comparing...")

        rowdf = self.df[self.df[self.config_col] == cfg]
        if rowdf.empty:
            messagebox.showerror("Error","No rows for config")
            return
        row = rowdf.iloc[0]

        # Detect payload columns
        headers = [h.lower() for h in self.df.columns]
        new_col = old_col = None

        for n in ["current","new","payload_json","new_payload"]:
            for h in headers:
                if n in h:
                    new_col = self.df.columns[headers.index(h)]
                    break
            if new_col: break

        for n in ["old","previous","prev_payload"]:
            for h in headers:
                if n in h:
                    old_col = self.df.columns[headers.index(h)]
                    break
            if old_col: break

        if not new_col or not old_col:
            messagebox.showerror("Error","Could not detect OLD/NEW columns")
            return

        old = safe_parse(row[old_col])
        new = safe_parse(row[new_col])

        self.show_json(old,new)
        self.compute_diff(old,new)

    # ------------------------------
    # JSON Display
    # ------------------------------
    def show_json(self, old, new):
        self.txt_old.delete("1.0", tk.END)
        self.txt_new.delete("1.0", tk.END)

        try: old_j = json.dumps(old, indent=4)
        except: old_j = str(old)
        try: new_j = json.dumps(new, indent=4)
        except: new_j = str(new)

        self.txt_old.insert(tk.END, old_j)
        self.txt_new.insert(tk.END, new_j)

    # ------------------------------
    # Diff Engine
    # ------------------------------
    def compute_diff(self, old, new):
        self.tree.delete(*self.tree.get_children())
        diff = DeepDiff(old,new,ignore_order=True,verbose_level=2)

        count=0
        if "values_changed" in diff:
            for path, d in diff["values_changed"].items():
                self.tree.insert("", "end", values=(path,"changed", d["old_value"], d["new_value"]))
                count+=1
        if "dictionary_item_added" in diff:
            for path in diff["dictionary_item_added"]:
                self.tree.insert("", "end", values=(path,"added","","ADDED"))
                count+=1
        if "dictionary_item_removed" in diff:
            for path in diff["dictionary_item_removed"]:
                self.tree.insert("", "end", values=(path,"removed","REMOVED",""))
                count+=1

        self.status.config(text=f"Total changes: {count}")


# ============================================================
# Run App
# ============================================================
if __name__ == "__main__":
    app = GeminiUltraApp()
    app.mainloop()
